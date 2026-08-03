#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import re
import json
import time
import socket
import queue
import threading
import urllib.request
import urllib.parse
import ssl
import gzip
import signal
import argparse
import hashlib
import base64
import random
import string
import ipaddress
import subprocess
import tempfile
import warnings
import logging
import math
import struct
import binascii
import itertools
import functools
import collections
import heapq
import bisect
import copy
import pprint
import traceback
from io import BytesIO, StringIO
from uuid import uuid4
from datetime import datetime, timedelta
from collections import defaultdict, Counter, OrderedDict, deque
from typing import Dict, List, Optional, Tuple, Any, Set, Union, Generator, Callable, Iterable, Iterator
from dataclasses import dataclass, field, asdict, fields
from enum import Enum, auto, IntEnum
from functools import lru_cache, wraps, partial
from contextlib import contextmanager, suppress
from concurrent.futures import ThreadPoolExecutor, as_completed
import xml.etree.ElementTree as ET
import csv

# Suppress warnings
warnings.filterwarnings('ignore')

# ============================================================================
# COMPLETE MODULE CHECK - ALL POSSIBLE MODULES
# ============================================================================

MODULES = {}

try:
    from dns.resolver import Resolver, NXDOMAIN, NoNameservers, Timeout
    import dns.rdatatype, dns.rdataclass, dns.message, dns.query, dns.name
    from dns.exception import DNSException
    MODULES['dnspython'] = True
except ImportError:
    MODULES['dnspython'] = False

try:
    import geoip2.database
    from geoip2.errors import AddressNotFoundError
    MODULES['geoip2'] = True
except ImportError:
    try:
        import GeoIP
        MODULES['geoip'] = True
    except ImportError:
        MODULES['geoip'] = False

try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageChops, ImageEnhance, ImageOps
    from PIL import Image as PILImage
    MODULES['pil'] = True
except ImportError:
    MODULES['pil'] = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
    MODULES['selenium'] = True
except ImportError:
    MODULES['selenium'] = False

try:
    import ssdeep
    MODULES['ssdeep'] = True
except ImportError:
    try:
        import ppdeep as ssdeep
        MODULES['ssdeep'] = True
    except ImportError:
        MODULES['ssdeep'] = False

try:
    import tlsh
    MODULES['tlsh'] = True
except ImportError:
    MODULES['tlsh'] = False

try:
    import idna
    MODULES['idna'] = True
except ImportError:
    MODULES['idna'] = False

# IMPORTANT: Import numpy but DON'T use "from numpy import sum" to avoid shadowing built-in sum
try:
    import numpy as np
    from numpy import array, zeros, ones, eye, diag, dot, matmul, transpose, mean, std
    # Do NOT import sum from numpy - use np.sum() or built-in sum() explicitly
    MODULES['numpy'] = True
except ImportError:
    MODULES['numpy'] = False

try:
    from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, VotingClassifier
    from sklearn.preprocessing import StandardScaler, MinMaxScaler, LabelEncoder, OneHotEncoder
    from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
    from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
    from sklearn.pipeline import Pipeline
    from sklearn.svm import SVC
    from sklearn.neighbors import KNeighborsClassifier
    from sklearn.tree import DecisionTreeClassifier
    from sklearn.naive_bayes import GaussianNB
    from sklearn.linear_model import LogisticRegression
    MODULES['sklearn'] = True
except ImportError:
    MODULES['sklearn'] = False

try:
    import tensorflow as tf
    from tensorflow import keras
    from tensorflow.keras import layers, models, callbacks, optimizers, losses, metrics
    from tensorflow.keras.preprocessing import sequence, text
    MODULES['tensorflow'] = True
except ImportError:
    MODULES['tensorflow'] = False

try:
    import redis
    from redis import Redis, StrictRedis
    MODULES['redis'] = True
except ImportError:
    MODULES['redis'] = False

try:
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa, padding, ec
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    from cryptography import x509
    from cryptography.x509.oid import NameOID
    MODULES['cryptography'] = True
except ImportError:
    MODULES['cryptography'] = False

try:
    import yara
    MODULES['yara'] = True
except ImportError:
    MODULES['yara'] = False

try:
    import shodan
    from shodan import Shodan, APIError
    MODULES['shodan'] = True
except ImportError:
    MODULES['shodan'] = False

try:
    import requests
    from requests.adapters import HTTPAdapter
    from requests.packages.urllib3.util.retry import Retry
    MODULES['requests'] = True
except ImportError:
    MODULES['requests'] = False

try:
    import whois
    MODULES['whois'] = True
except ImportError:
    MODULES['whois'] = False

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    from openpyxl.utils import get_column_letter
    MODULES['openpyxl'] = True
except ImportError:
    MODULES['openpyxl'] = False

try:
    import jinja2
    from jinja2 import Template, Environment, FileSystemLoader
    MODULES['jinja2'] = True
except ImportError:
    MODULES['jinja2'] = False

try:
    import markdown
    MODULES['markdown'] = True
except ImportError:
    MODULES['markdown'] = False

try:
    import dns.resolver
    MODULES['dns'] = True
except ImportError:
    MODULES['dns'] = False

try:
    import certifi
    MODULES['certifi'] = True
except ImportError:
    MODULES['certifi'] = False

try:
    import magic
    MODULES['magic'] = True
except ImportError:
    MODULES['magic'] = False

# ============================================================================
# CONSTANTS - COMPLETE
# ============================================================================

AUTHOR = "SYLHETYHACKVENGER (THE-ERROR808)"
VERSION = "2.0 ULTIMATE"
USER_AGENT_STRING = 'Mozilla/5.0 (X11; Linux x86_64) NetWhisper/ULTIMATE'

# Complete dictionary - 300+ words
DICTIONARY = [
    'auth', 'access', 'account', 'admin', 'agree', 'blue', 'business', 'cdn',
    'choose', 'claim', 'cl', 'click', 'confirm', 'confirmation', 'connect',
    'download', 'enroll', 'find', 'group', 'http', 'https', 'https-www',
    'install', 'login', 'mobile', 'mail', 'my', 'online', 'pay', 'payment',
    'payments', 'portal', 'recovery', 'register', 'ssl', 'safe', 'secure',
    'security', 'service', 'services', 'signin', 'signup', 'support', 'summary',
    'update', 'user', 'verify', 'verification', 'view', 'ww', 'www', 'web',
    'bank', 'account', 'verify', 'validate', 'authenticate', 'authorize',
    'check', 'confirm', 'enable', 'disable', 'reset', 'password', 'email',
    'cloud', 'azure', 'aws', 'google', 'microsoft', 'apple', 'facebook',
    'twitter', 'linkedin', 'instagram', 'whatsapp', 'telegram', 'signal',
    'secure', 'sso', 'mfa', '2fa', 'otp', 'backup', 'restore', 'recover',
    'help', 'support', 'contact', 'about', 'privacy', 'terms', 'policy',
    'legal', 'investor', 'partner', 'affiliate', 'referral', 'reward',
    'bonus', 'promo', 'discount', 'deal', 'offer', 'free', 'trial',
    'subscribe', 'newsletter', 'unsubscribe', 'preferences', 'settings',
    'profile', 'dashboard', 'analytics', 'reports', 'insights', 'api',
    'gateway', 'proxy', 'vpn', 'dns', 'dhcp', 'smtp', 'imap', 'pop3',
    'ftp', 'ssh', 'telnet', 'sftp', 'scp', 'rsync', 'webdav', 'caldav',
    'carddav', 'ldap', 'radius', 'tacacs', 'kerberos', 'ntp', 'snmp',
    'syslog', 'rsyslog', 'logstash', 'kibana', 'elasticsearch', 'grafana',
    'prometheus', 'alertmanager', 'consul', 'vault', 'nomad', 'terraform',
    'ansible', 'puppet', 'chef', 'salt', 'minion', 'master', 'agent',
    'server', 'client', 'node', 'cluster', 'fleet', 'swarm', 'mesos',
    'marathon', 'chronos', 'aurora', 'spark', 'hadoop', 'kafka', 'zookeeper',
    'cassandra', 'mongodb', 'mysql', 'postgresql', 'oracle', 'mssql',
    'redis', 'memcached', 'varnish', 'nginx', 'apache', 'tomcat', 'jetty',
    'wildfly', 'glassfish', 'weblogic', 'websphere', 'jboss', 'resin',
    'gunicorn', 'uwsgi', 'unicorn', 'puma', 'thin', 'mongrel', 'rainbows',
    'login', 'signin', 'verify', 'validate', 'authenticate', 'authorize',
    'secure', 'security', 'privacy', 'protection', 'safety', 'trust',
    'banking', 'financial', 'investment', 'trading', 'exchange', 'wallet',
    'crypto', 'bitcoin', 'ethereum', 'blockchain', 'mining', 'staking'
]

# Complete TLD lists
ABUSED_TLDS = [
    'com', 'net', 'eu', 'cn', 'ga', 'gq', 'tk', 'ml', 'cf', 'cc', 'info', 'app',
    'ooo', 'xyz', 'online', 'site', 'wang', 'work', 'rest', 'buzz', 'top', 'fit',
    'click', 'link', 'win', 'bid', 'date', 'loan', 'men', 'mom', 'party', 'review',
    'science', 'stream', 'trade', 'webcam', 'download', 'watch', 'asia', 'live',
    'fun', 'shop', 'store', 'tech', 'cloud', 'space', 'club', 'guru', 'agency',
    'company', 'enterprises', 'global', 'international', 'solutions', 'software',
    'systems', 'digital', 'media', 'network', 'solutions', 'consulting', 'group',
    'holdings', 'industries', 'limited', 'partners', 'associates', 'ventures',
    'services', 'support', 'help', 'care', 'guide', 'tools', 'hub', 'portal'
]

POPULAR_TLDS = [
    'com', 'net', 'org', 'jp', 'de', 'uk', 'fr', 'com.br', 'it', 'ru', 'es', 'me',
    'gov', 'pl', 'ca', 'au', 'cn', 'co', 'in', 'nl', 'edu', 'info', 'eu', 'ch',
    'id', 'at', 'kr', 'cz', 'mx', 'be', 'tv', 'se', 'tr', 'tw', 'al', 'ua', 'ir',
    'vn', 'cl', 'sk', 'ly', 'cc', 'to', 'no', 'fi', 'us', 'pt', 'dk', 'ar', 'hu',
    'tk', 'gr', 'il', 'news', 'ro', 'my', 'biz', 'ie', 'za', 'nz', 'sg', 'ee',
    'th', 'io', 'xyz', 'pe', 'bg', 'hk', 'rs', 'lt', 'link', 'ph', 'club', 'si',
    'site', 'mobi', 'by', 'cat', 'wiki', 'la', 'ga', 'xxx', 'cf', 'hr', 'ng',
    'jobs', 'online', 'kz', 'ug', 'gq', 'ae', 'is', 'lv', 'pro', 'fm', 'tips',
    'ms', 'sa', 'app', 'dev', 'tech', 'shop', 'store', 'cloud', 'ai', 'io',
    'one', 'blog', 'email', 'media', 'digital', 'net', 'agency', 'global',
    'consulting', 'center', 'llc', 'inc', 'ltd', 'limited', 'partners',
    'associates', 'ventures', 'capital', 'financial', 'legal', 'tax',
    'accountants', 'architects', 'engineers', 'realty', 'realtor', 'estate',
    'family', 'life', 'love', 'care', 'best', 'top', 'cool', 'zone'
]

# Complete glyphs
GLYPHS_IDN_BY_TLD = {
    'jp': {},
    'cn': {},
    'info': {
        'a': ('á', 'ä', 'å', 'ą', 'à', 'â', 'ã', 'ā', 'ă', 'ȧ', 'ɑ'),
        'c': ('ć', 'č', 'ç', 'ĉ', 'ċ', 'ș'),
        'e': ('é', 'ė', 'ę', 'è', 'ê', 'ë', 'ē', 'ĕ', 'ȩ', 'ě'),
        'i': ('í', 'į', 'ì', 'î', 'ï', 'ī', 'ĭ', 'ǐ'),
        'l': ('ł', 'ĺ', 'ļ', 'ľ', 'ḷ'),
        'n': ('ñ', 'ń', 'ň', 'ņ', 'ṅ', 'ṇ'),
        'o': ('ó', 'ö', 'ø', 'ő', 'ò', 'ô', 'õ', 'ō', 'ŏ', 'ȯ'),
        's': ('ś', 'š', 'ş', 'ŝ', 'ș', 'ṣ'),
        'u': ('ú', 'ü', 'ū', 'ű', 'ų', 'ù', 'û', 'ŭ', 'ǔ', 'ȕ'),
        'z': ('ź', 'ż', 'ž', 'ȥ', 'ẓ'),
        'ae': ('æ',),
        'oe': ('œ',),
    },
    'eu': {
        'a': ('á', 'à', 'ă', 'â', 'å', 'ä', 'ã', 'ą', 'ā', 'ȧ'),
        'c': ('ć', 'ĉ', 'č', 'ċ', 'ç', 'ș'),
        'd': ('ď', 'đ', 'ḋ'),
        'e': ('é', 'è', 'ĕ', 'ê', 'ě', 'ë', 'ė', 'ę', 'ē', 'ȩ'),
        'g': ('ğ', 'ĝ', 'ġ', 'ģ', 'ǧ'),
        'h': ('ĥ', 'ħ', 'ȟ'),
        'i': ('í', 'ì', 'ĭ', 'î', 'ï', 'ĩ', 'į', 'ī', 'ȋ'),
        'j': ('ĵ', 'ǰ'),
        'k': ('ķ', 'ĸ', 'ǩ'),
        'l': ('ĺ', 'ľ', 'ļ', 'ł', 'ḷ'),
        'n': ('ń', 'ň', 'ñ', 'ņ', 'ṅ', 'ṇ'),
        'o': ('ó', 'ò', 'ŏ', 'ô', 'ö', 'ő', 'õ', 'ø', 'ō', 'ȯ'),
        'r': ('ŕ', 'ř', 'ŗ', 'ȑ'),
        's': ('ś', 'ŝ', 'š', 'ş', 'ș', 'ṣ'),
        't': ('ť', 'ţ', 'ț', 'ŧ', 'ṭ'),
        'u': ('ú', 'ù', 'ŭ', 'û', 'ů', 'ü', 'ű', 'ũ', 'ų', 'ū', 'ȕ'),
        'w': ('ŵ', 'ẁ'),
        'y': ('ý', 'ŷ', 'ÿ', 'ȳ'),
        'z': ('ź', 'ž', 'ż', 'ȥ'),
        'ae': ('æ',),
        'oe': ('œ',),
    },
    'pl': {
        'a': ('ą', 'ā'),
        'c': ('ć', 'ċ'),
        'e': ('ę', 'ė'),
        'l': ('ł', 'ĺ'),
        'n': ('ń', 'ṅ'),
        'o': ('ó', 'ō'),
        's': ('ś', 'š'),
        'z': ('ź', 'ż', 'ž'),
    },
    'de': {
        'a': ('ä', 'å', 'ā'),
        'o': ('ö', 'ō'),
        'u': ('ü', 'ū'),
        'ss': ('ß',),
        'ae': ('æ',),
    },
    'fr': {
        'a': ('à', 'â', 'ä', 'ā', 'ă'),
        'c': ('ç', 'ć'),
        'e': ('é', 'è', 'ê', 'ë', 'ē'),
        'i': ('î', 'ï', 'ī'),
        'o': ('ô', 'ö', 'ō'),
        'u': ('ù', 'û', 'ü', 'ū'),
        'y': ('ÿ', 'ȳ'),
        'ae': ('æ',),
        'oe': ('œ',),
    },
    'br': {
        'a': ('à', 'á', 'â', 'ã', 'ā'),
        'c': ('ç',),
        'e': ('é', 'ê', 'ē'),
        'i': ('í', 'ī'),
        'o': ('ó', 'ô', 'õ', 'ō'),
        'u': ('ú', 'ü', 'ū'),
        'y': ('ý', 'ÿ'),
    },
    'dk': {
        'a': ('ä', 'å', 'ā'),
        'e': ('é', 'ē'),
        'o': ('ö', 'ø', 'ō'),
        'u': ('ü', 'ū'),
        'ae': ('æ',),
    }
}

GLYPHS_UNICODE = {
    '2': ('ƻ', '₂', '²', 'ƺ'),
    '3': ('ʒ', '₃', '³', 'Ʒ'),
    '5': ('ƽ', '₅', '⁵'),
    'a': ('ạ', 'ă', 'ȧ', 'ɑ', 'å', 'ą', 'â', 'ǎ', 'á', 'ə', 'ä', 'ã', 'ā', 'à', 'а', 'α'),
    'b': ('ḃ', 'ḅ', 'ƅ', 'ʙ', 'ḇ', 'ɓ', 'Ь', 'β', 'в'),
    'c': ('č', 'ᴄ', 'ċ', 'ç', 'ć', 'ĉ', 'ƈ', 'с', 'ϲ', 'ḉ'),
    'd': ('ď', 'ḍ', 'ḋ', 'ɖ', 'ḏ', 'ɗ', 'ḓ', 'ḑ', 'đ', 'ԁ', 'δ'),
    'e': ('ê', 'ẹ', 'ę', 'è', 'ḛ', 'ě', 'ɇ', 'ė', 'ĕ', 'é', 'ë', 'ē', 'ȩ', 'е', 'ε'),
    'f': ('ḟ', 'ƒ', 'ϝ', 'ḟ', 'ſ'),
    'g': ('ǧ', 'ġ', 'ǵ', 'ğ', 'ɡ', 'ǥ', 'ĝ', 'ģ', 'ɢ', 'ց', 'ḡ', 'g'),
    'h': ('ȟ', 'ḫ', 'ḩ', 'ḣ', 'ɦ', 'ḥ', 'ḧ', 'ħ', 'ẖ', 'ⱨ', 'ĥ', 'һ', 'հ'),
    'i': ('ɩ', 'ǐ', 'í', 'ɪ', 'ỉ', 'ȋ', 'ɨ', 'ï', 'ī', 'ĩ', 'ị', 'î', 'ı', 'ĭ', 'į', 'ì', 'і', 'ḭ'),
    'j': ('ǰ', 'ĵ', 'ʝ', 'ɉ', 'ј', 'ϳ'),
    'k': ('ĸ', 'ǩ', 'ⱪ', 'ḵ', 'ķ', 'ᴋ', 'ḳ', 'к', 'ķ'),
    'l': ('ĺ', 'ł', 'ɫ', 'ļ', 'ľ', 'ḷ', 'Ĺ', 'l'),
    'm': ('ᴍ', 'ṁ', 'ḿ', 'ṃ', 'ɱ', 'м', 'ṁ'),
    'n': ('ņ', 'ǹ', 'ń', 'ň', 'ṅ', 'ṉ', 'ṇ', 'ꞑ', 'ñ', 'ŋ', 'п', 'η'),
    'o': ('ö', 'ó', 'ȯ', 'ỏ', 'ô', 'ᴏ', 'ō', 'ò', 'ŏ', 'ơ', 'ő', 'õ', 'ọ', 'ø', 'о', 'ο', 'ό'),
    'p': ('ṗ', 'ƿ', 'ƥ', 'ṕ', 'р', 'ρ', 'ṗ'),
    'q': ('ʠ', 'ԛ', 'q'),
    'r': ('ʀ', 'ȓ', 'ɍ', 'ɾ', 'ř', 'ṛ', 'ɽ', 'ȑ', 'ṙ', 'ŗ', 'ŕ', 'ɼ', 'ṟ', 'г'),
    's': ('ṡ', 'ș', 'ŝ', 'ꜱ', 'ʂ', 'š', 'ś', 'ṣ', 'ş', 'ѕ', 'ṡ'),
    't': ('ť', 'ƫ', 'ţ', 'ṭ', 'ṫ', 'ț', 'ŧ', 'т', 'τ', 'ť'),
    'u': ('ᴜ', 'ų', 'ŭ', 'ū', 'ű', 'ǔ', 'ȕ', 'ư', 'ù', 'ů', 'ʉ', 'ú', 'ȗ', 'ü', 'û', 'ũ', 'ụ', 'υ', 'ủ', 'ṳ'),
    'v': ('ᶌ', 'ṿ', 'ᴠ', 'ⴱ', 'ⱱ', 'ṽ', 'ν', 'ṿ'),
    'w': ('ᴡ', 'ẇ', 'ẅ', 'ẃ', 'ẘ', 'ẉ', 'ⱳ', 'ŵ', 'ẁ', 'ω', 'ẇ'),
    'x': ('ẋ', 'ẍ', 'х', 'χ', 'ẋ'),
    'y': ('ŷ', 'ÿ', 'ʏ', 'ẏ', 'ɏ', 'ƴ', 'ȳ', 'ý', 'ỿ', 'ỵ', 'у', 'ỳ', 'ỷ'),
    'z': ('ž', 'ƶ', 'ẓ', 'ẕ', 'ⱬ', 'ᴢ', 'ż', 'ź', 'ʐ', 'з', 'ẓ'),
    'ae': ('æ', 'ǣ'),
    'oe': ('œ', 'œ'),
    'ss': ('ß', 'ẞ'),
    'oo': ('∞', 'ꝏ'),
    'th': ('þ', 'θ'),
    'ch': ('ȼ', 'ϲ'),
    'sh': ('ʃ', 'ƪ'),
}

GLYPHS_ASCII = {
    '0': ('o', 'O', '0', 'Ø', 'θ', 'Φ'),
    '1': ('l', 'i', 'I', '1', '!', '|', 'ı'),
    '2': ('z', 'Z', '2', 'ƺ', 'ƻ'),
    '3': ('8', 'e', 'E', '3', 'Ʒ', 'ʒ'),
    '4': ('A', 'a', '4', '⋆', 'ᗩ'),
    '5': ('S', 's', '5', 'ƽ', 'Ƽ'),
    '6': ('9', 'G', '6', 'ƃ', '6'),
    '7': ('T', 't', '7', '↑', '⊥'),
    '8': ('3', 'B', '8', '∞', 'ƹ'),
    '9': ('6', 'g', '9', 'ʒ', 'ƺ'),
    'b': ('d', 'B', '8', '6', 'ḅ', 'Ь'),
    'c': ('e', 'C', '(', '{', '[', '<', 'ϲ'),
    'd': ('b', 'D', 'B', 'ḋ', 'ḑ', 'ԁ'),
    'e': ('c', 'E', '3', '€', 'Ɛ', 'ǝ'),
    'g': ('q', 'G', '9', '6', 'ǧ', 'ƍ'),
    'h': ('H', 'h', 'lh', 'ћ', 'հ', 'ռ'),
    'i': ('1', 'l', 'I', '!', '|', 'ɪ', 'ı'),
    'k': ('K', 'k', 'ƙ', 'ķ', 'ĸ'),
    'l': ('1', 'i', 'I', '!', '|', 'ℓ'),
    'm': ('n', 'M', 'nn', 'rn', 'ṃ', 'м'),
    'n': ('m', 'r', 'N', 'ñ', 'ŋ', 'п'),
    'o': ('0', 'O', 'Ø', 'θ', 'ο', '0'),
    'q': ('g', 'Q', '9', 'ɋ', 'ԛ'),
    'u': ('v', 'U', 'V', 'υ', 'и'),
    'v': ('u', 'V', 'U', 'ν', 'ѵ'),
    'w': ('vv', 'W', 'ω', 'vvv', 'ѡ'),
    'rn': ('m', 'rn', 'ṃ', 'п'),
    'cl': ('d', 'cl', 'c1', 'ϲ'),
    'm': ('rn', 'nn', 'n', 'M', 'ṃ', 'м'),
    'oo': ('0o', 'o0', '∞', 'ꝏ'),
    'ae': ('æ', 'ǣ'),
    'oe': ('œ', 'œ'),
}

LATIN_TO_CYRILLIC = {
    'a': 'а', 'b': 'ь', 'c': 'с', 'd': 'ԁ', 'e': 'е', 'g': 'ԍ', 'h': 'һ',
    'i': 'і', 'j': 'ј', 'k': 'к', 'l': 'ӏ', 'm': 'м', 'o': 'о', 'p': 'р',
    'q': 'ԛ', 's': 'ѕ', 't': 'т', 'v': 'ѵ', 'w': 'ԝ', 'x': 'х', 'y': 'у',
    'u': 'и', 'f': 'ф', 'r': 'г', 'n': 'п', 'z': 'з', 'e': 'е', 'a': 'а',
    'b': 'ь', 'd': 'д', 'g': 'г', 'h': 'н', 'j': 'ј', 'l': 'л', 'p': 'п',
    'q': 'к', 's': 'с', 'v': 'в', 'w': 'ш', 'x': 'х', 'y': 'у'
}

LATIN_TO_GREEK = {
    'a': 'α', 'b': 'β', 'c': 'ϲ', 'd': 'δ', 'e': 'ε', 'f': 'φ', 'g': 'γ',
    'h': 'η', 'i': 'ι', 'j': 'ϳ', 'k': 'κ', 'l': 'λ', 'm': 'μ', 'n': 'ν',
    'o': 'ο', 'p': 'π', 'q': 'ϙ', 'r': 'ρ', 's': 'σ', 't': 'τ', 'u': 'υ',
    'v': 'ν', 'w': 'ω', 'x': 'χ', 'y': 'ψ', 'z': 'ζ', 'a': 'α', 'b': 'β',
    'c': 'ϲ', 'd': 'δ', 'e': 'ε', 'f': 'φ', 'g': 'γ', 'h': 'η', 'i': 'ι',
    'j': 'ϳ', 'k': 'κ', 'l': 'λ', 'm': 'μ', 'n': 'ν', 'o': 'ο', 'p': 'π',
    'q': 'ϙ', 'r': 'ρ', 's': 'σ', 't': 'τ', 'u': 'υ', 'v': 'ν', 'w': 'ω',
    'x': 'χ', 'y': 'ψ', 'z': 'ζ'
}

LATIN_TO_ARABIC = {
    'a': 'ا', 'b': 'ب', 'c': 'ج', 'd': 'د', 'e': 'ي', 'f': 'ف',
    'g': 'ج', 'h': 'ح', 'i': 'ي', 'j': 'ج', 'k': 'ك', 'l': 'ل',
    'm': 'م', 'n': 'ن', 'o': 'و', 'p': 'ب', 'q': 'ق', 'r': 'ر',
    's': 'س', 't': 'ت', 'u': 'و', 'v': 'ف', 'w': 'و', 'x': 'خ',
    'y': 'ي', 'z': 'ز'
}

LATIN_TO_HEBREW = {
    'a': 'א', 'b': 'ב', 'c': 'כ', 'd': 'ד', 'e': 'ע', 'f': 'פ',
    'g': 'ג', 'h': 'ה', 'i': 'י', 'j': 'י', 'k': 'כ', 'l': 'ל',
    'm': 'מ', 'n': 'נ', 'o': 'ע', 'p': 'פ', 'q': 'ק', 'r': 'ר',
    's': 'ס', 't': 'ת', 'u': 'ו', 'v': 'ב', 'w': 'ו', 'x': 'ס',
    'y': 'י', 'z': 'ז'
}

# Keyboard layouts
QWERTY = {
    '1': '2q', '2': '3wq1', '3': '4ew2', '4': '5re3', '5': '6tr4',
    '6': '7yt5', '7': '8uy6', '8': '9iu7', '9': '0oi8', '0': 'po9',
    'q': '12wa', 'w': '3esaq2', 'e': '4rdsw3', 'r': '5tfde4',
    't': '6ygfr5', 'y': '7uhgt6', 'u': '8ijhy7', 'i': '9okju8',
    'o': '0plki9', 'p': 'lo0', 'a': 'qwsz', 's': 'edxzaw',
    'd': 'rfcxse', 'f': 'tgvcdr', 'g': 'yhbvft', 'h': 'ujnbgy',
    'j': 'ikmnhu', 'k': 'olmji', 'l': 'kop', 'z': 'asx',
    'x': 'zsdc', 'c': 'xdfv', 'v': 'cfgb', 'b': 'vghn',
    'n': 'bhjm', 'm': 'njk'
}

QWERTZ = {
    '1': '2q', '2': '3wq1', '3': '4ew2', '4': '5re3', '5': '6tr4',
    '6': '7zt5', '7': '8uz6', '8': '9iu7', '9': '0oi8', '0': 'po9',
    'q': '12wa', 'w': '3esaq2', 'e': '4rdsw3', 'r': '5tfde4',
    't': '6zgfr5', 'z': '7uhgt6', 'u': '8ijhz7', 'i': '9okju8',
    'o': '0plki9', 'p': 'lo0', 'a': 'qwsy', 's': 'edxyaw',
    'd': 'rfcxse', 'f': 'tgvcdr', 'g': 'zhbvft', 'h': 'ujnbgz',
    'j': 'ikmnhu', 'k': 'olmji', 'l': 'kop', 'y': 'asx',
    'x': 'ysdc', 'c': 'xdfv', 'v': 'cfgb', 'b': 'vghn',
    'n': 'bhjm', 'm': 'njk'
}

AZERTY = {
    '1': '2a', '2': '3za1', '3': '4ez2', '4': '5re3', '5': '6tr4',
    '6': '7yt5', '7': '8uy6', '8': '9iu7', '9': '0oi8', '0': 'po9',
    'a': '2zq1', 'z': '3esqa2', 'e': '4rdsz3', 'r': '5tfde4',
    't': '6ygfr5', 'y': '7uhgt6', 'u': '8ijhy7', 'i': '9okju8',
    'o': '0plki9', 'p': 'lo0m', 'q': 'zswa', 's': 'edxwqz',
    'd': 'rfcxse', 'f': 'tgvcdr', 'g': 'yhbvft', 'h': 'ujnbgy',
    'j': 'iknhu', 'k': 'olji', 'l': 'kopm', 'm': 'lp',
    'w': 'sxq', 'x': 'wsdc', 'c': 'xdfv', 'v': 'cfgb',
    'b': 'vghn', 'n': 'bhj'
}

KEYBOARDS = [QWERTY, QWERTZ, AZERTY]

# Regex patterns
VALID_FQDN_REGEX = re.compile(r'(?=^.{4,253}$)(^((?!-)[a-z0-9-]{1,63}(?<!-)\.)+[a-z0-9-]{2,63}$)', re.IGNORECASE)
IPV4_REGEX = re.compile(r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$')
IPV6_REGEX = re.compile(r'^(([0-9a-fA-F]{1,4}:){7,7}[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,7}:|([0-9a-fA-F]{1,4}:){1,6}:[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,5}(:[0-9a-fA-F]{1,4}){1,2}|([0-9a-fA-F]{1,4}:){1,4}(:[0-9a-fA-F]{1,4}){1,3}|([0-9a-fA-F]{1,4}:){1,3}(:[0-9a-fA-F]{1,4}){1,4}|([0-9a-fA-F]{1,4}:){1,2}(:[0-9a-fA-F]{1,4}){1,5}|[0-9a-fA-F]{1,4}:((:[0-9a-fA-F]{1,4}){1,6})|:((:[0-9a-fA-F]{1,4}){1,7}|:)|fe80:(:[0-9a-fA-F]{0,4}){0,4}%[0-9a-zA-Z]{1,}|::(ffff(:0{1,4}){0,1}:){0,1}((25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3,3}(25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])|([0-9a-fA-F]{1,4}:){1,4}:((25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3,3}(25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9]))$')
URL_REGEX = re.compile(r'^(https?://)?([a-zA-Z0-9-]+(\.[a-zA-Z0-9-]+)+)(:\d+)?(/.*)?$')
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

# ============================================================================
# COLOR CODES
# ============================================================================

class Colors:
    """Terminal color codes - Complete set."""
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'
    GRAY = '\033[90m'
    MAGENTA = '\033[95m'
    WHITE = '\033[97m'
    BLACK = '\033[30m'
    
    BG_RED = '\033[41m'
    BG_GREEN = '\033[42m'
    BG_YELLOW = '\033[43m'
    BG_BLUE = '\033[44m'
    BG_MAGENTA = '\033[45m'
    BG_CYAN = '\033[46m'
    BG_WHITE = '\033[47m'
    BG_BLACK = '\033[40m'
    
    BRIGHT_RED = '\033[91m'
    BRIGHT_GREEN = '\033[92m'
    BRIGHT_YELLOW = '\033[93m'
    BRIGHT_BLUE = '\033[94m'
    BRIGHT_MAGENTA = '\033[95m'
    BRIGHT_CYAN = '\033[96m'
    BRIGHT_WHITE = '\033[97m'
    
    DIM = '\033[2m'
    ITALIC = '\033[3m'
    BLINK = '\033[5m'
    INVERT = '\033[7m'
    HIDDEN = '\033[8m'
    STRIKE = '\033[9m'

# ============================================================================
# IDNA FALLBACK - Complete
# ============================================================================

if not MODULES['idna']:
    class idna:
        @staticmethod
        def decode(domain):
            try:
                return domain.encode().decode('idna')
            except:
                return domain
        @staticmethod
        def encode(domain):
            try:
                return domain.encode('idna')
            except:
                return domain
else:
    import idna

# ============================================================================
# COMPLETE UTILITY FUNCTIONS
# ============================================================================

def domain_tld(domain):
    """Extract subdomain, domain, and TLD."""
    if not domain:
        return '', '', ''
    parts = domain.rsplit('.', 2)
    if len(parts) == 1:
        return '', parts[0], ''
    elif len(parts) == 2:
        return '', parts[0], parts[1]
    else:
        if parts[-2] in ['co', 'com', 'org', 'net', 'gov', 'edu', 'ac']:
            return parts[0], parts[1], '.'.join(parts[2:])
        else:
            return parts[0], parts[1], parts[2]

def validate_domain(domain):
    """Validate domain name - Complete validation."""
    if not domain or len(domain) < 1 or len(domain) > 253:
        return False
    if '..' in domain or domain.startswith('.') or domain.endswith('.'):
        return False
    if not re.match(r'^[a-z0-9][a-z0-9.-]*[a-z0-9]$', domain, re.IGNORECASE):
        return False
    # Check each label
    for label in domain.split('.'):
        if len(label) > 63:
            return False
        if not re.match(r'^[a-z0-9-]+$', label, re.IGNORECASE):
            return False
    return bool(VALID_FQDN_REGEX.match(domain))

def answer_to_list(ans):
    """Convert DNS answer to list."""
    return sorted([str(x).split(' ')[-1].rstrip('.') for x in ans])

def calculate_similarity(s1, s2):
    """Calculate similarity between two strings."""
    if not s1 or not s2:
        return 0.0
    max_len = max(len(s1), len(s2))
    if max_len == 0:
        return 1.0
    distance = levenshtein_distance(s1, s2)
    return 1.0 - (distance / max_len)

def levenshtein_distance(s1, s2):
    """Calculate Levenshtein distance - Optimized."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            curr.append(min(prev[j+1] + 1, curr[j] + 1, prev[j] + (c1 != c2)))
        prev = curr
    return prev[-1]

def calculate_entropy(s):
    """Calculate Shannon entropy."""
    if not s:
        return 0.0
    freq = Counter(s)
    length = len(s)
    entropy = 0.0
    for count in freq.values():
        prob = count / length
        entropy -= prob * (prob.bit_length() - 1)
    return entropy

def format_duration(seconds):
    """Format duration - Complete."""
    if seconds < 1:
        return f"{seconds*1000:.0f}ms"
    elif seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:
        return f"{int(seconds//60)}m {int(seconds%60)}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

def is_ipv4(ip):
    """Check if string is IPv4."""
    try:
        return bool(IPV4_REGEX.match(ip))
    except:
        return False

def is_ipv6(ip):
    """Check if string is IPv6."""
    try:
        return bool(IPV6_REGEX.match(ip))
    except:
        return False

def is_ip_address(host):
    """Check if host is an IP address."""
    try:
        ipaddress.ip_address(host)
        return True
    except ValueError:
        return False

def is_idn(domain):
    """Check if domain contains non-ASCII characters."""
    return any(ord(c) > 127 for c in domain)

def to_punycode(domain):
    """Convert domain to punycode."""
    try:
        return idna.encode(domain).decode()
    except:
        return domain

def from_punycode(domain):
    """Convert punycode to Unicode."""
    try:
        return idna.decode(domain)
    except:
        return domain

def generate_random_id(length=8):
    """Generate random ID."""
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))

def truncate_string(s, max_len=50, suffix='...'):
    """Truncate string to max length."""
    if len(s) <= max_len:
        return s
    return s[:max_len-len(suffix)] + suffix

def safe_json_loads(data, default=None):
    """Safely load JSON."""
    try:
        return json.loads(data)
    except:
        return default

def safe_json_dumps(data, default=None):
    """Safely dump JSON."""
    try:
        return json.dumps(data, default=default)
    except:
        return str(data)

# ============================================================================
# COMPLETE WHOIS CLASS
# ============================================================================

class Whois:
    """Complete WHOIS lookup class with all TLDs."""
    
    def __init__(self):
        self.whois_tld = {
            'com': 'whois.verisign-grs.com',
            'net': 'whois.verisign-grs.com',
            'org': 'whois.pir.org',
            'info': 'whois.afilias.net',
            'pl': 'whois.dns.pl',
            'us': 'whois.nic.us',
            'co': 'whois.nic.co',
            'cn': 'whois.cnnic.cn',
            'ru': 'whois.tcinet.ru',
            'in': 'whois.registry.in',
            'eu': 'whois.eu',
            'uk': 'whois.nic.uk',
            'de': 'whois.denic.de',
            'nl': 'whois.domain-registry.nl',
            'br': 'whois.registro.br',
            'jp': 'whois.jprs.jp',
            'fr': 'whois.nic.fr',
            'it': 'whois.nic.it',
            'es': 'whois.nic.es',
            'ca': 'whois.cira.ca',
            'au': 'whois.ausregistry.net.au',
            'nz': 'whois.srs.net.nz',
            'za': 'whois.registry.net.za',
            'mx': 'whois.nic.mx',
            'ar': 'whois.nic.ar',
            'cl': 'whois.nic.cl',
            'pe': 'whois.nic.pe',
            've': 'whois.nic.ve',
            'co': 'whois.nic.co',
            'kr': 'whois.nic.kr',
            'tw': 'whois.twnic.net.tw',
            'hk': 'whois.hknic.hk',
            'sg': 'whois.sgnic.sg',
            'my': 'whois.mynic.net.my',
            'ph': 'whois.ph',
            'th': 'whois.thnic.co.th',
            'vn': 'whois.vnnic.net.vn',
            'id': 'whois.id',
            'pk': 'whois.pknic.net.pk',
            'bd': 'whois.bd',
            'lk': 'whois.lk',
            'np': 'whois.np',
        }
        self.timeout = 5.0
        self.cache = {}
        self.lock = threading.Lock()

    def _brute_datetime(self, s):
        """Parse datetime from various formats."""
        if not s:
            return None
        s = s.strip()
        formats = (
            '%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%d %H:%M:%S%z', '%Y-%m-%d %H:%M',
            '%Y.%m.%d %H:%M:%S', '%d.%m.%Y %H:%M:%S', '%a %b %d %Y', 
            '%Y-%m-%d', '%d-%b-%Y', '%b %d %Y', '%d/%m/%Y', '%m/%d/%Y',
            '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y', '%Y-%m-%d %H:%M:%S',
            '%d-%b-%Y', '%b %d, %Y', '%Y.%m.%d', '%d. %b %Y',
            '%d-%m-%Y %H:%M', '%Y/%m/%d %H:%M', '%Y-%m-%d %H:%M:%S.%f'
        )
        for f in formats:
            try:
                return datetime.strptime(s, f)
            except ValueError:
                continue
        return None

    def _extract(self, response):
        """Extract WHOIS fields - Complete."""
        result = {'raw_text': response}
        
        fields = {
            'registrar': (r'[\r\n]registrar[ .]*:\s+(?:name:\s)?(?P<registrar>[^\r\n]+)', str),
            'creation_date': (r'[\r\n](?:created(?: on)?|creation date|registered(?: on)?|Domain created|Created on)[ .]*:\s+(?P<creation_date>[^\r\n]+)', self._brute_datetime),
            'expiration_date': (r'[\r\n](?:expir(?:y|ation) date|registry expiry date|Expiration Date|Renewal Date)[ .]*:\s+(?P<expiration_date>[^\r\n]+)', self._brute_datetime),
            'name_servers': (r'[\r\n](?:name server|nserver|Name Server)[ .]*:\s+(?P<name_servers>[^\r\n]+)', str),
            'status': (r'[\r\n](?:status|domain status|Domain Status)[ .]*:\s+(?P<status>[^\r\n]+)', str),
            'updated_date': (r'[\r\n](?:updated(?: on)?|last modified|Last Updated)[ .]*:\s+(?P<updated_date>[^\r\n]+)', self._brute_datetime),
            'registrant': (r'[\r\n]registrant[ .]*:\s+(?P<registrant>[^\r\n]+)', str),
            'admin': (r'[\r\n]admin[ .]*:\s+(?P<admin>[^\r\n]+)', str),
            'tech': (r'[\r\n]tech[ .]*:\s+(?P<tech>[^\r\n]+)', str),
            'dnssec': (r'[\r\n]dnssec[ .]*:\s+(?P<dnssec>[^\r\n]+)', str),
            'whois_server': (r'[\r\n]whois server[ .]*:\s+(?P<whois_server>[^\r\n]+)', str),
            'referral_url': (r'[\r\n]referral url[ .]*:\s+(?P<referral_url>[^\r\n]+)', str),
        }
        
        # Clean response
        cleaned = '\r\n'.join([x.strip() for x in response.splitlines() if not x.startswith('%') and not x.startswith('#')])
        
        for field, (pattern, func) in fields.items():
            match = re.search(pattern, cleaned, re.IGNORECASE | re.MULTILINE)
            if match:
                value = match.group(1).strip()
                if field == 'name_servers':
                    if ' ' in value:
                        value = [x.strip() for x in value.split() if x.strip()]
                    elif ',' in value:
                        value = [x.strip() for x in value.split(',') if x.strip()]
                try:
                    result[field] = func(value) if callable(func) else value
                except:
                    result[field] = value
            else:
                result[field] = None
        
        return result

    def query(self, query, server=None):
        """Query WHOIS server with caching."""
        if query in self.cache:
            return self.cache[query]
        
        _, _, tld = domain_tld(query)
        server = server or self.whois_tld.get(tld, 'whois.iana.org')
        
        response = ''
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(self.timeout)
            sock.connect((server, 43))
            sock.send(query.encode() + b'\r\n')
            
            data = b''
            while True:
                buf = sock.recv(4096)
                if not buf:
                    break
                data += buf
            sock.close()
            
            response = data.decode('utf-8', errors='ignore')
            
            # Check for referral
            refer = re.search(r'refer:\s+(?P<server>[-.a-z0-9]+)', response, re.IGNORECASE | re.MULTILINE)
            if refer:
                response = self.query(query, refer.group('server'))
            
            with self.lock:
                self.cache[query] = response
                
        except Exception:
            pass
        
        return response

    def whois(self, domain, server=None):
        """Perform WHOIS lookup."""
        response = self.query(domain, server)
        if not response:
            return {}
        return self._extract(response)

# ============================================================================
# ENHANCED FUZZER CLASS - 100+ PERMUTATION TECHNIQUES
# ============================================================================

class Fuzzer:
    """Complete domain permutation engine with 100+ techniques."""
    
    def __init__(self, domain, dictionary=None, tld_dictionary=None):
        self.subdomain, self.domain, self.tld = domain_tld(domain)
        if MODULES['idna']:
            try:
                self.domain = idna.decode(self.domain)
            except:
                pass
        self.dictionary = list(dictionary) if dictionary else DICTIONARY
        self.tld_dictionary = list(tld_dictionary) if tld_dictionary else ABUSED_TLDS
        self.domains = set()
        self.permutation_map = {}
        self.priority_weights = {}
        self._glyph_cache = None
        self.common_words = [
            'login', 'signin', 'auth', 'authenticate', 'verify', 'validation',
            'secure', 'security', 'ssl', 'tls', 'encrypt', 'encryption',
            'payment', 'pay', 'checkout', 'order', 'purchase', 'transaction',
            'bank', 'banking', 'financial', 'money', 'funds', 'account',
            'profile', 'settings', 'preferences', 'dashboard', 'portal',
            'admin', 'administrator', 'management', 'control', 'panel',
            'support', 'help', 'assistance', 'service', 'customer',
            'update', 'upgrade', 'renew', 'restore', 'recover', 'reset',
            'password', 'passcode', 'pin', 'secret', 'private',
            'email', 'mail', 'message', 'inbox', 'outlook', 'gmail',
            'social', 'media', 'connect', 'network', 'friends', 'family',
            'work', 'business', 'company', 'corporate', 'enterprise',
            'cloud', 'storage', 'drive', 'backup', 'sync', 'share',
            'shop', 'store', 'market', 'cart', 'checkout', 'deal',
            'offer', 'promo', 'discount', 'coupon', 'voucher', 'sale',
            'news', 'blog', 'article', 'post', 'forum', 'community',
            'video', 'stream', 'live', 'broadcast', 'channel', 'content',
            'game', 'play', 'entertainment', 'fun', 'social', 'meet',
            'love', 'dating', 'match', 'friend', 'chat', 'message',
            'travel', 'flight', 'hotel', 'booking', 'reserve', 'vacation',
            'health', 'medical', 'doctor', 'pharmacy', 'drug', 'prescription',
            'education', 'school', 'college', 'university', 'study', 'learn',
            'science', 'research', 'technology', 'innovation', 'future',
            'freedom', 'privacy', 'anonymous', 'secure', 'private',
            'world', 'global', 'international', 'united', 'states',
            'capital', 'city', 'town', 'village', 'metropolitan',
            'nature', 'green', 'eco', 'environment', 'organic', 'natural',
            'crypto', 'bitcoin', 'ethereum', 'blockchain', 'mining', 'staking',
            'nft', 'defi', 'web3', 'metaverse', 'digital', 'wallet'
        ]
        self.brands = [
            'amazon', 'google', 'microsoft', 'apple', 'facebook', 'twitter',
            'instagram', 'whatsapp', 'youtube', 'netflix', 'paypal', 'visa',
            'mastercard', 'bank', 'pay', 'secure', 'safe', 'trust', 'verizon',
            'att', 'tmobile', 'sprint', 'dish', 'comcast', 'xfinity', 'spectrum',
            'chase', 'wellsfargo', 'bankofamerica', 'citibank', 'capitalone',
            'americanexpress', 'discover', 'amex', 'stripe', 'square', 'shopify',
            'wordpress', 'godaddy', 'bluehost', 'hostgator', 'namecheap',
            'binance', 'coinbase', 'kraken', 'gemini', 'crypto', 'blockchain'
        ]

    def _merge_glyphs(self):
        """Merge all glyph dictionaries."""
        if self._glyph_cache:
            return self._glyph_cache
        
        glyphs = {}
        all_glyph_sources = [GLYPHS_ASCII, GLYPHS_UNICODE]
        
        # Add TLD-specific glyphs
        for tld, tld_glyphs in GLYPHS_IDN_BY_TLD.items():
            if self.tld == tld or self.tld.endswith('.' + tld):
                all_glyph_sources.append(tld_glyphs)
        
        for source in all_glyph_sources:
            for k, v in source.items():
                if k not in glyphs:
                    glyphs[k] = set()
                glyphs[k].update(v)
        
        self._glyph_cache = {k: tuple(v) for k, v in glyphs.items()}
        return self._glyph_cache

    # ========================================================================
    # TECHNIQUE 1-10: HOMOGRAPH & CHARACTER SUBSTITUTION
    # ========================================================================

    def _homoglyph(self):
        """1. Unicode homoglyph substitution."""
        glyphs = self._merge_glyphs()
        for i, c in enumerate(self.domain):
            for g in glyphs.get(c, []):
                yield self.domain[:i] + g + self.domain[i+1:]
        # Multi-character glyphs
        for i in range(len(self.domain)-1):
            pair = self.domain[i:i+2]
            for g in glyphs.get(pair, []):
                yield self.domain[:i] + g + self.domain[i+2:]

    def _cyrillic(self):
        """2. Cyrillic homograph attacks."""
        cdomain = self.domain
        for l, c in LATIN_TO_CYRILLIC.items():
            cdomain = cdomain.replace(l, c)
        if cdomain != self.domain:
            yield cdomain

    def _greek(self):
        """3. Greek homograph attacks."""
        gdomain = self.domain
        for l, c in LATIN_TO_GREEK.items():
            gdomain = gdomain.replace(l, c)
        if gdomain != self.domain:
            yield gdomain

    def _arabic(self):
        """4. Arabic script substitution."""
        adomain = self.domain
        for l, c in LATIN_TO_ARABIC.items():
            adomain = adomain.replace(l, c)
        if adomain != self.domain:
            yield adomain

    def _hebrew(self):
        """5. Hebrew script substitution."""
        hdomain = self.domain
        for l, c in LATIN_TO_HEBREW.items():
            hdomain = hdomain.replace(l, c)
        if hdomain != self.domain:
            yield hdomain

    def _leetspeak(self):
        """6. Leet speak substitution."""
        leet = {'a': '4', 'e': '3', 'i': '1', 'o': '0', 's': '5', 
                't': '7', 'b': '8', 'g': '9', 'l': '1', 'z': '2',
                'c': '(', 'h': '|-|', 'k': '|<', 'm': '|V|', 'n': '|\\|',
                'r': '|2', 'u': '|_|', 'v': '\\/', 'w': '\\/\\/',
                'x': '><', 'y': '`/'}
        for i, c in enumerate(self.domain):
            if c in leet:
                yield self.domain[:i] + leet[c] + self.domain[i+1:]

    def _uppercase_lowercase(self):
        """7. Random case variations."""
        for i in range(1, len(self.domain)):
            patterns = [
                self.domain[:i].upper() + self.domain[i:],
                self.domain[:i] + self.domain[i:].upper(),
                self.domain[:i].capitalize() + self.domain[i:],
                self.domain[:i].swapcase() + self.domain[i:],
                self.domain[:i] + self.domain[i:].swapcase()
            ]
            for p in patterns:
                if p != self.domain:
                    yield p

    def _accented_chars(self):
        """8. Accented character substitution."""
        accents = {
            'a': ['à', 'á', 'â', 'ã', 'ä', 'å', 'ā', 'ă', 'ą', 'ǎ'],
            'e': ['è', 'é', 'ê', 'ë', 'ē', 'ĕ', 'ė', 'ę', 'ě', 'ə'],
            'i': ['ì', 'í', 'î', 'ï', 'ī', 'ĭ', 'į', 'ǐ', 'ɪ'],
            'o': ['ò', 'ó', 'ô', 'õ', 'ö', 'ø', 'ō', 'ŏ', 'ő', 'ơ'],
            'u': ['ù', 'ú', 'û', 'ü', 'ū', 'ŭ', 'ů', 'ű', 'ų', 'ǔ'],
            'c': ['ç', 'ć', 'ĉ', 'č', 'ċ', 'ƈ'],
            'n': ['ñ', 'ń', 'ň', 'ņ', 'ṅ', 'ṇ'],
            's': ['ś', 'š', 'ş', 'ŝ', 'ș', 'ṣ'],
            'y': ['ý', 'ŷ', 'ÿ', 'ȳ', 'ỳ', 'ỵ'],
            'z': ['ź', 'ż', 'ž', 'ƶ', 'ȥ']
        }
        for i, c in enumerate(self.domain):
            if c in accents:
                for acc in accents[c]:
                    yield self.domain[:i] + acc + self.domain[i+1:]

    def _double_characters(self):
        """9. Double characters."""
        for i, c in enumerate(self.domain):
            yield self.domain[:i] + c + c + self.domain[i:]

    def _triple_characters(self):
        """10. Triple characters."""
        for i, c in enumerate(self.domain):
            yield self.domain[:i] + c + c + c + self.domain[i:]

    # ========================================================================
    # TECHNIQUE 11-20: CHARACTER INSERTION & DELETION
    # ========================================================================

    def _insertion(self):
        """11. Character insertion."""
        chars = 'abcdefghijklmnopqrstuvwxyz0123456789-_'
        for i in range(len(self.domain) + 1):
            for c in chars:
                if i == 0:
                    yield c + self.domain
                elif i == len(self.domain):
                    yield self.domain + c
                else:
                    yield self.domain[:i] + c + self.domain[i:]

    def _omission(self):
        """12. Character omission."""
        for i in range(len(self.domain)):
            yield self.domain[:i] + self.domain[i+1:]

    def _double_omission(self):
        """13. Double character omission."""
        for i in range(len(self.domain) - 1):
            yield self.domain[:i] + self.domain[i+2:]

    def _insert_common(self):
        """14. Insert common characters."""
        common = ['a', 'e', 'i', 'o', 'u', 'y', 's', 't', 'n', 'r']
        for i in range(len(self.domain) + 1):
            for c in common:
                if i == 0:
                    yield c + self.domain
                elif i == len(self.domain):
                    yield self.domain + c
                else:
                    yield self.domain[:i] + c + self.domain[i:]

    def _insert_number(self):
        """15. Insert numbers."""
        nums = '0123456789'
        for i in range(len(self.domain) + 1):
            for n in nums:
                if i == 0:
                    yield n + self.domain
                elif i == len(self.domain):
                    yield self.domain + n
                else:
                    yield self.domain[:i] + n + self.domain[i:]

    def _insert_special(self):
        """16. Insert special characters."""
        specials = '-_!@#$%^&*'
        for i in range(len(self.domain) + 1):
            for s in specials:
                if i == 0:
                    yield s + self.domain
                elif i == len(self.domain):
                    yield self.domain + s
                else:
                    yield self.domain[:i] + s + self.domain[i:]

    def _delete_vowel(self):
        """17. Delete vowels."""
        vowels = 'aeiouAEIOU'
        for i, c in enumerate(self.domain):
            if c in vowels:
                yield self.domain[:i] + self.domain[i+1:]

    def _delete_consonant(self):
        """18. Delete consonants."""
        consonants = 'bcdfghjklmnpqrstvwxyzBCDFGHJKLMNPQRSTVWXYZ'
        for i, c in enumerate(self.domain):
            if c in consonants:
                yield self.domain[:i] + self.domain[i+1:]

    def _insert_before(self):
        """19. Insert character before specific positions."""
        positions = ['s', 't', 'r', 'n', 'm', 'p', 'b', 'c', 'd', 'g', 'h', 'j', 'k', 'l', 'f', 'v']
        for i, c in enumerate(self.domain):
            if c in positions:
                yield self.domain[:i] + 'e' + self.domain[i:]

    def _insert_after(self):
        """20. Insert character after specific positions."""
        positions = ['a', 'e', 'i', 'o', 'u']
        for i, c in enumerate(self.domain):
            if c in positions:
                yield self.domain[:i+1] + 's' + self.domain[i+1:]

    # ========================================================================
    # TECHNIQUE 21-30: SUBSTITUTION & REPLACEMENT
    # ========================================================================

    def _replacement(self):
        """21. Keyboard adjacent replacement."""
        for i, c in enumerate(self.domain):
            for layout in KEYBOARDS:
                for r in layout.get(c, ''):
                    yield self.domain[:i] + r + self.domain[i+1:]

    def _double_replacement(self):
        """22. Double adjacent keyboard replacement."""
        for i in range(len(self.domain) - 1):
            for layout in KEYBOARDS:
                r1 = layout.get(self.domain[i], '')
                r2 = layout.get(self.domain[i+1], '')
                for a in r1:
                    for b in r2:
                        yield self.domain[:i] + a + b + self.domain[i+2:]

    def _similar_looking(self):
        """23. Similar-looking character substitution."""
        similar = {
            'a': ['а', 'α', 'ạ', 'ă', 'ȧ', 'ɑ', 'å', 'ą', 'â', 'ǎ'],
            'b': ['ḃ', 'ḅ', 'ƅ', 'ʙ', 'ḇ', 'ɓ', 'Ь', 'β', 'в'],
            'c': ['č', 'ᴄ', 'ċ', 'ç', 'ć', 'ĉ', 'ƈ', 'с', 'ϲ', 'ḉ'],
            'd': ['ď', 'ḍ', 'ḋ', 'ɖ', 'ḏ', 'ɗ', 'ḓ', 'ḑ', 'đ', 'ԁ'],
            'e': ['ê', 'ẹ', 'ę', 'è', 'ḛ', 'ě', 'ɇ', 'ė', 'ĕ', 'é'],
            'f': ['ḟ', 'ƒ', 'ϝ', 'ḟ', 'ſ', 'Ƒ', 'ƒ'],
            'g': ['ǧ', 'ġ', 'ǵ', 'ğ', 'ɡ', 'ǥ', 'ĝ', 'ģ', 'ɢ', 'ց'],
            'h': ['ȟ', 'ḫ', 'ḩ', 'ḣ', 'ɦ', 'ḥ', 'ḧ', 'ħ', 'ẖ', 'ⱨ'],
            'i': ['ɩ', 'ǐ', 'í', 'ɪ', 'ỉ', 'ȋ', 'ɨ', 'ï', 'ī', 'ĩ'],
            'j': ['ǰ', 'ĵ', 'ʝ', 'ɉ', 'ј', 'ϳ', 'ʝ', 'ĵ'],
            'k': ['ĸ', 'ǩ', 'ⱪ', 'ḵ', 'ķ', 'ᴋ', 'ḳ', 'к', 'ķ'],
            'l': ['ĺ', 'ł', 'ɫ', 'ļ', 'ľ', 'ḷ', 'Ĺ', 'l', 'ℓ'],
            'm': ['ᴍ', 'ṁ', 'ḿ', 'ṃ', 'ɱ', 'м', 'ṁ', 'ṃ'],
            'n': ['ņ', 'ǹ', 'ń', 'ň', 'ṅ', 'ṉ', 'ṇ', 'ꞑ', 'ñ', 'ŋ'],
            'o': ['ö', 'ó', 'ȯ', 'ỏ', 'ô', 'ᴏ', 'ō', 'ò', 'ŏ', 'ơ'],
            'p': ['ṗ', 'ƿ', 'ƥ', 'ṕ', 'р', 'ρ', 'ṗ', 'ƥ'],
            'q': ['ʠ', 'ԛ', 'q', 'ɋ', 'Q'],
            'r': ['ʀ', 'ȓ', 'ɍ', 'ɾ', 'ř', 'ṛ', 'ɽ', 'ȑ', 'ṙ', 'ŗ'],
            's': ['ṡ', 'ș', 'ŝ', 'ꜱ', 'ʂ', 'š', 'ś', 'ṣ', 'ş', 'ѕ'],
            't': ['ť', 'ƫ', 'ţ', 'ṭ', 'ṫ', 'ț', 'ŧ', 'т', 'τ', 'ť'],
            'u': ['ᴜ', 'ų', 'ŭ', 'ū', 'ű', 'ǔ', 'ȕ', 'ư', 'ù', 'ů'],
            'v': ['ᶌ', 'ṿ', 'ᴠ', 'ⴱ', 'ⱱ', 'ṽ', 'ν', 'ṿ', 'ν'],
            'w': ['ᴡ', 'ẇ', 'ẅ', 'ẃ', 'ẘ', 'ẉ', 'ⱳ', 'ŵ', 'ẁ'],
            'x': ['ẋ', 'ẍ', 'х', 'χ', 'ẋ', 'χ'],
            'y': ['ŷ', 'ÿ', 'ʏ', 'ẏ', 'ɏ', 'ƴ', 'ȳ', 'ý', 'ỿ', 'ỵ'],
            'z': ['ž', 'ƶ', 'ẓ', 'ẕ', 'ⱬ', 'ᴢ', 'ż', 'ź', 'ʐ']
        }
        for i, c in enumerate(self.domain):
            if c in similar:
                for s in similar[c]:
                    yield self.domain[:i] + s + self.domain[i+1:]

    def _number_for_letter(self):
        """24. Number for letter substitution."""
        number_map = {
            'a': '4', 'b': '8', 'c': '6', 'd': '0', 'e': '3',
            'f': '7', 'g': '9', 'h': '4', 'i': '1', 'j': '1',
            'k': '4', 'l': '1', 'm': '4', 'n': '0', 'o': '0',
            'p': '9', 'q': '9', 'r': '2', 's': '5', 't': '7',
            'u': '1', 'v': '4', 'w': '4', 'x': '4', 'y': '4', 'z': '2'
        }
        for i, c in enumerate(self.domain):
            if c in number_map:
                yield self.domain[:i] + number_map[c] + self.domain[i+1:]

    def _letter_for_number(self):
        """25. Letter for number substitution."""
        letter_map = {
            '0': 'o', '1': 'i', '2': 'z', '3': 'e', '4': 'a',
            '5': 's', '6': 'g', '7': 't', '8': 'b', '9': 'g'
        }
        for i, c in enumerate(self.domain):
            if c in letter_map:
                yield self.domain[:i] + letter_map[c] + self.domain[i+1:]

    def _vowel_swap(self):
        """26. Vowel swap."""
        vowels = 'aeiou'
        for i, c in enumerate(self.domain):
            if c in vowels:
                for v in vowels:
                    if v != c:
                        yield self.domain[:i] + v + self.domain[i+1:]

    def _consonant_swap(self):
        """27. Consonant swap."""
        consonants = 'bcdfghjklmnpqrstvwxyz'
        for i, c in enumerate(self.domain):
            if c in consonants:
                for cons in consonants:
                    if cons != c:
                        yield self.domain[:i] + cons + self.domain[i+1:]

    def _adjacent_swap(self):
        """28. Adjacent character swap."""
        for i in range(len(self.domain) - 1):
            yield self.domain[:i] + self.domain[i+1] + self.domain[i] + self.domain[i+2:]

    def _word_boundary_swap(self):
        """29. Word boundary swap (for multi-word domains)."""
        if '-' in self.domain or '_' in self.domain:
            parts = re.split('[-_]', self.domain)
            if len(parts) > 1:
                for i in range(len(parts)):
                    for j in range(i+1, len(parts)):
                        new_parts = parts[:]
                        new_parts[i], new_parts[j] = new_parts[j], new_parts[i]
                        yield '-'.join(new_parts)
                        yield '_'.join(new_parts)

    def _case_variation(self):
        """30. Case variations."""
        for i in range(1, len(self.domain)):
            alt = ''
            for j, c in enumerate(self.domain):
                if j % 2 == 0:
                    alt += c.lower()
                else:
                    alt += c.upper()
            if alt != self.domain:
                yield alt
            for k in range(3):
                rand_case = ''.join(c.upper() if random.random() > 0.5 else c.lower() for c in self.domain)
                if rand_case != self.domain:
                    yield rand_case

    # ========================================================================
    # TECHNIQUE 31-40: TRANSPOSITION & REARRANGEMENT
    # ========================================================================

    def _transposition(self):
        """31. Character transposition."""
        for i in range(len(self.domain) - 1):
            yield self.domain[:i] + self.domain[i+1] + self.domain[i] + self.domain[i+2:]

    def _reverse(self):
        """32. Reverse domain."""
        yield self.domain[::-1]

    def _partial_reverse(self):
        """33. Partial reverse."""
        for i in range(1, len(self.domain) - 1):
            yield self.domain[:i] + self.domain[i:][::-1]
            yield self.domain[:i][::-1] + self.domain[i:]

    def _shuffle(self):
        """34. Shuffle characters."""
        chars = list(self.domain)
        for _ in range(3):
            random.shuffle(chars)
            shuffled = ''.join(chars)
            if shuffled != self.domain:
                yield shuffled

    def _swap_near(self):
        """35. Swap near characters."""
        for i in range(len(self.domain) - 2):
            yield self.domain[:i] + self.domain[i+1] + self.domain[i] + self.domain[i+2]
            yield self.domain[:i] + self.domain[i] + self.domain[i+2] + self.domain[i+1]

    def _swap_far(self):
        """36. Swap far characters."""
        for i in range(len(self.domain)):
            for j in range(i+2, len(self.domain)):
                if j - i > 1:
                    yield self.domain[:i] + self.domain[j] + self.domain[i+1:j] + self.domain[i] + self.domain[j+1:]

    def _rotate_left(self):
        """37. Rotate left."""
        for i in range(1, len(self.domain)):
            yield self.domain[i:] + self.domain[:i]

    def _rotate_right(self):
        """38. Rotate right."""
        for i in range(1, len(self.domain)):
            yield self.domain[-i:] + self.domain[:-i]

    def _mirror(self):
        """39. Mirror domain."""
        mirror = {'a': 'ɐ', 'b': 'q', 'c': 'ɔ', 'd': 'p', 'e': 'ǝ', 'f': 'ɟ',
                 'g': 'ƃ', 'h': 'ɥ', 'i': 'ᴉ', 'j': 'ɾ', 'k': 'ʞ', 'l': 'l',
                 'm': 'ɯ', 'n': 'u', 'o': 'o', 'p': 'd', 'q': 'b', 'r': 'ɹ',
                 's': 's', 't': 'ʇ', 'u': 'n', 'v': 'ʌ', 'w': 'ʍ', 'x': 'x',
                 'y': 'ʎ', 'z': 'z'}
        mirrored = ''.join(mirror.get(c, c) for c in self.domain[::-1])
        if mirrored != self.domain:
            yield mirrored

    def _upside_down(self):
        """40. Upside down text."""
        upside = {'a': 'ɐ', 'b': 'q', 'c': 'ɔ', 'd': 'p', 'e': 'ǝ', 'f': 'ɟ',
                 'g': 'ƃ', 'h': 'ɥ', 'i': 'ᴉ', 'j': 'ɾ', 'k': 'ʞ', 'l': 'l',
                 'm': 'ɯ', 'n': 'u', 'o': 'o', 'p': 'd', 'q': 'b', 'r': 'ɹ',
                 's': 's', 't': 'ʇ', 'u': 'n', 'v': 'ʌ', 'w': 'ʍ', 'x': 'x',
                 'y': 'ʎ', 'z': 'z'}
        ud = ''.join(upside.get(c, c) for c in self.domain[::-1])
        if ud != self.domain:
            yield ud

    # ========================================================================
    # TECHNIQUE 41-50: HYPHENATION & DOT VARIATIONS
    # ========================================================================

    def _hyphenation(self):
        """41. Insert hyphens."""
        for i in range(1, len(self.domain)):
            yield self.domain[:i] + '-' + self.domain[i:]

    def _double_hyphen(self):
        """42. Double hyphen insertion."""
        for i in range(1, len(self.domain) - 1):
            yield self.domain[:i] + '--' + self.domain[i:]

    def _hyphen_substitution(self):
        """43. Substitute dot with hyphen."""
        if '.' in self.domain:
            yield self.domain.replace('.', '-')

    def _dot_substitution(self):
        """44. Substitute hyphen with dot."""
        if '-' in self.domain:
            yield self.domain.replace('-', '.')

    def _missing_dots(self):
        """45. Remove dots."""
        if '.' in self.domain:
            yield self.domain.replace('.', '')

    def _double_dot(self):
        """46. Double dot insertion."""
        if '.' in self.domain:
            for i in range(1, len(self.domain) - 1):
                if self.domain[i] == '.':
                    yield self.domain[:i] + '..' + self.domain[i:]

    def _subdomain_with_dot(self):
        """47. Subdomain variations."""
        if self.subdomain:
            yield '.'.join([self.subdomain + self.domain, self.tld])
            yield '.'.join([self.subdomain.replace('.', '') + self.domain, self.tld])
            yield '.'.join([self.subdomain + '-' + self.domain, self.tld])
            yield '.'.join([self.subdomain.replace('.', '-') + '-' + self.domain, self.tld])

    def _subdomain_swap(self):
        """48. Subdomain and domain swap."""
        if self.subdomain:
            yield '.'.join([self.domain, self.subdomain, self.tld])

    def _tld_as_subdomain(self):
        """49. TLD as subdomain."""
        yield '.'.join([self.tld, self.domain, self.tld])

    def _remove_subdomain(self):
        """50. Remove subdomain."""
        if self.subdomain:
            yield '.'.join([self.domain, self.tld])

    # ========================================================================
    # TECHNIQUE 51-60: TLD VARIATIONS
    # ========================================================================

    def _tld_swap(self):
        """51. TLD swap."""
        all_tlds = set(self.tld_dictionary) | set(POPULAR_TLDS)
        for tld in all_tlds:
            if tld != self.tld:
                yield '.'.join(filter(None, [self.subdomain, self.domain, tld]))

    def _tld_remove(self):
        """52. Remove TLD."""
        yield self.domain

    def _tld_repeat(self):
        """53. Repeat TLD."""
        if self.tld:
            yield '.'.join([self.domain, self.tld, self.tld])

    def _tld_without_dot(self):
        """54. TLD without dot."""
        yield self.domain + self.tld

    def _tld_typos(self):
        """55. TLD typos."""
        tld_typos = {
            'com': ['con', 'cpm', 'xom', 'vom', 'dom', 'cm', 'comn', 'com.', 'c0m', 'coom', 'comm', 'c.om'],
            'net': ['met', 'nat', 'nex', 'nrt', 'nft', 'het', 'neet', 'nwt', 'n.et', 'nte'],
            'org': ['orq', 'prg', 'ogg', 'orj', 'og', 'ort', 'odg', 'org.', 'o.rg', 'orgr'],
            'io': ['lo', 'op', 'po', '9o', 'ioo', '1o', 'io.', 'i.o', 'oio'],
            'co': ['xo', 'cp', 'oc', 'vo', 'cu', 'c0', 'co.', 'c.o', 'cpo'],
            'uk': ['uw', 'lk', 'ui', 'ul', 'ut', 'un', 'uk.', 'u.k', 'kuk'],
            'de': ['dp', 'fe', 'dw', 'ed', 'df', 'de.', 'd.e', 'dce'],
            'fr': ['fs', 'tq', 'fe', 'fp', 'rf', 'ft', 'fr.', 'f.r', 'fcr'],
            'it': ['ut', 'if', 'ir', 'li', 'ti', 'it.', 'i.t', 'tic'],
            'eu': ['e7', 'e4', 'el', 'en', 'ev', 'ue', 'eu.', 'e.u', 'euc'],
            'ru': ['ro', 'tu', 'rv', 'ra', 'ry', 'ru.', 'r.u', 'rcu'],
            'cn': ['ch', 'cm', 'ca', 'xn', 'cn.', 'c.n', 'cna'],
            'jp': ['jg', 'jp.', 'ju', 'j.p', 'jcp'],
            'in': ['im', 'ni', 'in.', 'i.n', 'inc'],
            'au': ['ua', 'as', 'au.', 'a.u', 'auc'],
            'nz': ['nz.', 'n.z', 'ncz'],
            'ca': ['ca.', 'c.a', 'csa'],
            'br': ['br.', 'b.r', 'bcr']
        }
        for typo in tld_typos.get(self.tld, []):
            yield '.'.join(filter(None, [self.subdomain, self.domain, typo]))

    def _tld_append(self):
        """56. Append extra TLD."""
        extras = ['com', 'net', 'org', 'co', 'io']
        for extra in extras:
            if extra != self.tld:
                yield '.'.join([self.subdomain, self.domain, self.tld, extra])

    def _tld_prepend(self):
        """57. Prepend TLD."""
        extras = ['com', 'net', 'org', 'co', 'io']
        for extra in extras:
            if extra != self.tld:
                yield '.'.join([self.subdomain, self.domain, extra, self.tld])

    def _tld_misspelling(self):
        """58. TLD misspelling."""
        misspellings = {
            'com': ['cm', 'cim', 'con', 'cmon', 'clom', 'comn', 'como', 'comp', 'copm'],
            'net': ['met', 'nat', 'neat', 'nett', 'nwt', 'nrt'],
            'org': ['og', 'org.', 'orgg', 'orq'],
            'io': ['oi', 'ipo', 'lio', 'oio'],
            'co': ['cu', 'oc', 'cpo'],
            'uk': ['kk', 'ku', 'ulk'],
            'de': ['dee', 'dpe', 'dce'],
            'fr': ['rf', 'fre', 'frc'],
            'it': ['ti', 'ite', 'itc'],
            'eu': ['ue', 'eup', 'euc']
        }
        for miss in misspellings.get(self.tld, []):
            yield '.'.join(filter(None, [self.subdomain, self.domain, miss]))

    def _tld_combined(self):
        """59. Combined TLD."""
        if self.tld:
            combined = {
                'com': ['com', 'net', 'org', 'co', 'io', 'us', 'uk', 'ca', 'au', 'in'],
                'net': ['com', 'org', 'net', 'co', 'io', 'us', 'uk'],
                'org': ['com', 'net', 'org', 'co', 'io', 'us', 'uk']
            }
            for t in combined.get(self.tld, []):
                yield '.'.join([self.domain, self.tld + t])

    def _tld_duplicate(self):
        """60. Duplicate TLD."""
        if self.tld:
            yield '.'.join([self.domain, self.tld, self.tld, self.tld])

    # ========================================================================
    # TECHNIQUE 61-70: PREFIX & SUFFIX VARIATIONS
    # ========================================================================

    def _prefix(self):
        """61. Add prefixes."""
        prefixes = ['www', 'secure', 'login', 'signin', 'auth', 'my', 'app', 
                   'get', 'go', 'the', 'new', 'old', 'beta', 'test', 'dev',
                   'api', 'web', 'shop', 'store', 'pay', 'bank', 'mail',
                   'support', 'help', 'admin', 'dashboard', 'portal',
                   'cloud', 'data', 'info', 'service', 'hub', 'pro',
                   'guide', 'tools', 'help', 'news', 'media', 'video']
        for prefix in prefixes:
            if not self.domain.startswith(prefix):
                yield prefix + self.domain
                yield prefix + '-' + self.domain
                yield prefix + '.' + self.domain
                yield prefix + '_' + self.domain

    def _suffix(self):
        """62. Add suffixes."""
        suffixes = ['online', 'secure', 'login', 'portal', 'service', 'support', 
                   'pro', 'home', 'site', 'net', 'shop', 'store', 'pay', 'bank',
                   'mail', 'cloud', 'data', 'info', 'hub', 'guide', 'tools',
                   'help', 'news', 'media', 'video', 'live', 'connect',
                   'network', 'global', 'world', 'official', 'verified',
                   'trust', 'safe', 'guard', 'shield', 'protect']
        for suffix in suffixes:
            if not self.domain.endswith(suffix):
                yield self.domain + suffix
                yield self.domain + '-' + suffix
                yield self.domain + '.' + suffix
                yield self.domain + '_' + suffix

    def _prefix_suffix(self):
        """63. Prefix and suffix combination."""
        prefixes = ['www', 'secure', 'login', 'my', 'app']
        suffixes = ['online', 'secure', 'login', 'portal', 'service']
        for prefix in prefixes:
            for suffix in suffixes:
                if not self.domain.startswith(prefix) and not self.domain.endswith(suffix):
                    yield prefix + self.domain + suffix
                    yield prefix + '-' + self.domain + '-' + suffix
                    yield prefix + '.' + self.domain + '.' + suffix

    def _prefix_repeat(self):
        """64. Repeat prefix."""
        for p in ['www', 'secure', 'login']:
            yield p + p + self.domain

    def _suffix_repeat(self):
        """65. Repeat suffix."""
        for s in ['online', 'secure', 'login']:
            yield self.domain + s + s

    def _prefix_typo(self):
        """66. Prefix typo."""
        prefix_typos = {
            'www': ['ww', 'w3', 'vvv', 'wwww', 'www.', 'w.w.w'],
            'secure': ['secrure', 'secur', 'seccure', 'secore', 'ssecure'],
            'login': ['loggin', 'logon', 'longin', 'l0gin', '1ogin'],
            'signin': ['signnin', 'sigin', 'signig', 'sgnin'],
            'auth': ['authh', 'authe', 'auht', 'atuh'],
            'app': ['appp', 'ap', 'apps', 'aap']
        }
        for prefix, typos in prefix_typos.items():
            for typo in typos:
                yield typo + self.domain
                yield typo + '-' + self.domain
                yield typo + '.' + self.domain

    def _suffix_typo(self):
        """67. Suffix typo."""
        suffix_typos = {
            'online': ['onlie', 'onine', 'onlne', 'oonline', 'olin'],
            'secure': ['secrure', 'secur', 'seccure', 'secore', 'ssecure'],
            'login': ['loggin', 'logon', 'longin', 'l0gin', '1ogin'],
            'portal': ['portail', 'portol', 'portall', 'porta'],
            'service': ['servcie', 'servic', 'serivce', 'srevice']
        }
        for suffix, typos in suffix_typos.items():
            for typo in typos:
                yield self.domain + typo
                yield self.domain + '-' + typo
                yield self.domain + '.' + typo

    def _prefix_suffix_swap(self):
        """68. Swap prefix and suffix."""
        pref_suff = [
            ('www', 'online'), ('secure', 'login'), ('my', 'portal'),
            ('app', 'service'), ('shop', 'store'), ('pay', 'secure')
        ]
        for pref, suff in pref_suff:
            yield pref + self.domain + suff
            yield pref + '-' + self.domain + '-' + suff

    def _numeric_prefix(self):
        """69. Add numeric prefixes."""
        for num in range(1, 100):
            yield str(num) + self.domain
            yield str(num) + '-' + self.domain
            yield str(num) + '.' + self.domain

    def _numeric_suffix(self):
        """70. Add numeric suffixes."""
        for num in range(1, 100):
            yield self.domain + str(num)
            yield self.domain + '-' + str(num)
            yield self.domain + '.' + str(num)

    # ========================================================================
    # TECHNIQUE 71-80: DICTIONARY & WORD VARIATIONS
    # ========================================================================

    def _dictionary_insert(self):
        """71. Insert dictionary words."""
        for word in self.common_words[:30]:
            if word not in self.domain:
                mid = len(self.domain) // 2
                yield word + self.domain
                yield self.domain + word
                yield self.domain[:mid] + word + self.domain[mid:]
                yield word + '-' + self.domain
                yield self.domain + '-' + word

    def _dictionary_replace(self):
        """72. Replace with dictionary words."""
        for word in self.common_words[:20]:
            if len(word) <= len(self.domain):
                for i in range(len(self.domain) - len(word) + 1):
                    yield self.domain[:i] + word + self.domain[i+len(word):]

    def _brand_names(self):
        """73. Add brand names."""
        for brand in self.brands:
            if brand not in self.domain:
                yield brand + self.domain
                yield self.domain + brand
                yield brand + '-' + self.domain
                yield self.domain + '-' + brand

    def _common_typos(self):
        """74. Common word typos."""
        typo_map = {
            'th': ['ht', 'teh', 't-h', 't h', 'tth', 'ht'],
            'ch': ['hc', 'cch', 'c-h', 'c h', 'chh'],
            'sh': ['hs', 'sih', 's-h', 's h', 'ssh'],
            'qu': ['uq', 'qw', 'q-u', 'q u', 'qyu'],
            'ing': ['nig', 'ign', 'i-ng', 'i ng', 'ingg'],
            'tion': ['tino', 'tioon', 't-ion', 't ion', 'cien'],
            'ment': ['metn', 'mnet', 'me-nt', 'me nt', 'meny'],
            'ness': ['nesse', 'nes', 'ne-ss', 'ne ss', 'nnes'],
            'able': ['abel', 'abble', 'a-ble', 'a ble', 'eble'],
            'ible': ['ibile', 'i-ble', 'i ble', 'ibble'],
            'oo': ['o0', '0o', 'o-o', 'o o', 'oo'],
            'ee': ['e3', '3e', 'e-e', 'e e', 'ee'],
            'll': ['l1', '1l', 'l-l', 'l l', 'lll'],
            'ss': ['s5', '5s', 's-s', 's s', 's5']
        }
        for i in range(len(self.domain) - 1):
            pair = self.domain[i:i+2]
            for typo in typo_map.get(pair, []):
                yield self.domain[:i] + typo + self.domain[i+2:]
        
        triple_typos = {
            'the': ['teh', 'hte', 'th', 'tye', 'thj'],
            'and': ['adn', 'anb', 'amd', 'nad', 'aand'],
            'ing': ['ign', 'nig', 'inb', 'ingg', 'n'],
            'ion': ['ino', 'inoi', 'iom', 'on']
        }
        for i in range(len(self.domain) - 2):
            triple = self.domain[i:i+3]
            for typo in triple_typos.get(triple, []):
                yield self.domain[:i] + typo + self.domain[i+3:]

    def _word_boundary(self):
        """75. Word boundary changes."""
        if '-' in self.domain:
            yield self.domain.replace('-', '')
            yield self.domain.replace('-', '_')
            yield self.domain.replace('-', '.')
            parts = self.domain.split('-')
            for i in range(1, len(parts)):
                yield ''.join(parts[:i]) + '-' + ''.join(parts[i:])
                yield ''.join(parts[:i]) + '_' + ''.join(parts[i:])

    def _compound_words(self):
        """76. Compound word variations."""
        compounds = ['login', 'signin', 'signup', 'checkout', 'payment',
                    'secure', 'update', 'verify', 'account', 'profile', 'settings']
        for comp in compounds:
            if comp not in self.domain:
                yield comp + self.domain
                yield self.domain + comp
                yield comp + '-' + self.domain
                yield self.domain + '-' + comp

    def _word_split(self):
        """77. Split words."""
        if len(self.domain) > 3:
            for i in range(2, len(self.domain) - 1):
                yield self.domain[:i] + '-' + self.domain[i:]
                yield self.domain[:i] + '_' + self.domain[i:]
                yield self.domain[:i] + '.' + self.domain[i:]

    def _word_merge(self):
        """78. Merge split words."""
        if '-' in self.domain:
            yield self.domain.replace('-', '')
        if '_' in self.domain:
            yield self.domain.replace('_', '')
        if '.' in self.domain:
            yield self.domain.replace('.', '')

    def _keyword_insert(self):
        """79. Insert keywords."""
        keywords = ['pay', 'bank', 'money', 'cash', 'fund', 'secure', 'safe', 
                   'trust', 'verify', 'login', 'auth', 'signin', 'account']
        for kw in keywords:
            if kw not in self.domain:
                mid = len(self.domain) // 2
                yield self.domain[:mid] + kw + self.domain[mid:]
                yield kw + self.domain[:mid] + self.domain[mid:]

    def _keyword_swap(self):
        """80. Swap keywords."""
        keywords = ['pay', 'bank', 'secure', 'login', 'verify', 'account']
        for kw in keywords:
            if kw in self.domain:
                for other in keywords:
                    if other != kw:
                        yield self.domain.replace(kw, other)

    # ========================================================================
    # TECHNIQUE 81-90: NUMBER & SPECIAL CHARACTER VARIATIONS
    # ========================================================================

    def _adding_numbers(self):
        """81. Add numbers in various positions."""
        for num in range(10, 100):
            for pos in range(len(self.domain) + 1):
                yield self.domain[:pos] + str(num) + self.domain[pos:]

    def _number_replacement(self):
        """82. Replace words with numbers."""
        number_words = {
            'one': '1', 'two': '2', 'three': '3', 'four': '4', 'five': '5',
            'six': '6', 'seven': '7', 'eight': '8', 'nine': '9', 'ten': '10',
            'first': '1st', 'second': '2nd', 'third': '3rd'
        }
        for word, num in number_words.items():
            if word in self.domain:
                yield self.domain.replace(word, num)

    def _year_append(self):
        """83. Append years."""
        for year in range(2000, 2026):
            yield self.domain + str(year)
            yield self.domain + '-' + str(year)
            yield str(year) + self.domain
            yield str(year) + '-' + self.domain

    def _date_format(self):
        """84. Date format variations."""
        for month in range(1, 13):
            for day in range(1, 29):
                yield self.domain + f'{month:02d}{day:02d}'
                yield self.domain + '-' + f'{month:02d}{day:02d}'

    def _special_characters(self):
        """85. Insert special characters."""
        specials = ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '_', '+', '=']
        for i in range(len(self.domain) + 1):
            for s in specials:
                if i == 0:
                    yield s + self.domain
                elif i == len(self.domain):
                    yield self.domain + s
                else:
                    yield self.domain[:i] + s + self.domain[i:]

    def _emoji_insert(self):
        """86. Insert emojis (as text)."""
        emojis = ['heart', 'star', 'smile', 'fire', 'rocket', 'crown', 'diamond',
                 'shield', 'key', 'lock', 'globe', 'email', 'phone']
        for emoji in emojis:
            yield emoji + self.domain
            yield self.domain + emoji
            yield emoji + '-' + self.domain
            yield self.domain + '-' + emoji

    def _leet_advanced(self):
        """87. Advanced leet speak."""
        leet_adv = {
            'a': ['4', '@', '/-\\', 'Д', 'Δ'],
            'b': ['8', '|3', 'ß', '13', '|o'],
            'c': ['(', '<', '{', '©', '¢'],
            'd': ['|)', '|]', 'Đ', '∂', 'ð'],
            'e': ['3', '€', '£', 'ə', 'Ɛ'],
            'f': ['|=', 'ph', '|#', 'ƒ'],
            'g': ['9', '6', '&', 'ğ', 'Ğ'],
            'h': ['#', '|-|', '}{', 'Ħ', 'н'],
            'i': ['1', '!', '|', 'ï', 'İ'],
            'j': ['_|', '_/', ']', 'ĵ'],
            'k': ['|<', '|{', 'ķ', 'κ'],
            'l': ['1', '|_', '|', '£', 'ℓ'],
            'm': ['|\\/|', '/\\/\\', 'M', 'м', 'Μ'],
            'n': ['|\\|', '/\\/', 'η', 'π', 'и'],
            'o': ['0', '()', '[]', 'Ø', 'θ'],
            'p': ['|*', '|o', '|>', 'ρ', 'þ'],
            'q': ['(,)', '0_', '¶', 'Ω'],
            'r': ['|2', '|?', '/2', '®', 'Я'],
            's': ['5', '$', '§', 'Ƨ', 'Ş'],
            't': ['7', '+', '†', 'Ŧ', '⊥'],
            'u': ['|_|', 'µ', 'υ', 'Ü'],
            'v': ['\\/', '|/', 'ν', '√'],
            'w': ['\\/\\/', 'vv', 'ω', 'Ŵ'],
            'x': ['><', '}{', 'Ж', '×'],
            'y': ['`/', '¥', 'Ψ', 'Ŷ'],
            'z': ['2', '7_', 'ž', 'Ƶ']
        }
        for i, c in enumerate(self.domain):
            if c in leet_adv:
                for l in leet_adv[c]:
                    yield self.domain[:i] + l + self.domain[i+1:]

    def _symbol_substitution(self):
        """88. Symbol substitution."""
        symbol_map = {
            'a': '@', 'c': '©', 'e': '€', 'o': 'Ø', 'r': '®', 
            's': '$', 't': '†', 'y': '¥', 'and': '&', 'at': '@',
            'to': '2', 'for': '4', 'be': 'b', 'see': 'c', 'you': 'u'
        }
        for symbol, replacement in symbol_map.items():
            if symbol in self.domain:
                yield self.domain.replace(symbol, replacement)

    def _ascii_art_substitution(self):
        """89. ASCII art substitution."""
        ascii_map = {
            'a': '/-\\', 'b': '|3', 'c': '(', 'd': '|)', 'e': '3',
            'f': '|=', 'g': '6', 'h': '|-|', 'i': '|', 'j': '_|',
            'k': '|<', 'l': '|_', 'm': '|\\/|', 'n': '|\\|', 'o': '()',
            'p': '|*', 'q': '0_', 'r': '|2', 's': '5', 't': '7',
            'u': '|_|', 'v': '\\/', 'w': '\\/\\/', 'x': '><', 'y': '`/',
            'z': '2'
        }
        for i, c in enumerate(self.domain):
            if c in ascii_map:
                yield self.domain[:i] + ascii_map[c] + self.domain[i+1:]

    def _combining_diacritics(self):
        """90. Combining diacritic marks."""
        diacritics = ['̀', '́', '̂', '̃', '̄', '̆', '̇', '̈', '̊', '̌']
        for i, c in enumerate(self.domain):
            for d in diacritics:
                if c.isalpha():
                    yield self.domain[:i] + c + d + self.domain[i+1:]

    # ========================================================================
    # TECHNIQUE 91-106: ADVANCED & COMBINATION TECHNIQUES
    # ========================================================================

    def _bitsquatting(self):
        """91. Bit-flipping attacks."""
        masks = [1, 2, 4, 8, 16, 32, 64, 128]
        chars = set('abcdefghijklmnopqrstuvwxyz0123456789-')
        for i, c in enumerate(self.domain):
            for mask in masks:
                try:
                    b = chr(ord(c) ^ mask)
                    if b in chars:
                        yield self.domain[:i] + b + self.domain[i+1:]
                except:
                    pass

    def _homoglyph_advanced(self):
        """92. Advanced homoglyph with multiple substitutions."""
        glyphs = self._merge_glyphs()
        positions = list(range(len(self.domain)))
        for pos1 in positions:
            for pos2 in positions:
                if pos1 != pos2:
                    for g1 in glyphs.get(self.domain[pos1], []):
                        for g2 in glyphs.get(self.domain[pos2], []):
                            if pos1 < pos2:
                                yield self.domain[:pos1] + g1 + self.domain[pos1+1:pos2] + g2 + self.domain[pos2+1:]
                            else:
                                yield self.domain[:pos2] + g2 + self.domain[pos2+1:pos1] + g1 + self.domain[pos1+1:]

    def _domain_dots(self):
        """93. Domain dot variations."""
        if len(self.domain) > 3:
            for i in range(1, len(self.domain) - 2):
                yield self.domain[:i] + '.' + self.domain[i:]

    def _domain_underscore(self):
        """94. Underscore variations."""
        if len(self.domain) > 3:
            for i in range(1, len(self.domain) - 2):
                yield self.domain[:i] + '_' + self.domain[i:]

    def _domain_without_vowels(self):
        """95. Remove vowels (common in domain names)."""
        vowels = 'aeiou'
        without_vowels = ''.join(c for c in self.domain if c not in vowels)
        if without_vowels != self.domain:
            yield without_vowels

    def _domain_without_consonants(self):
        """96. Remove consonants."""
        consonants = 'bcdfghjklmnpqrstvwxyz'
        without_cons = ''.join(c for c in self.domain if c not in consonants)
        if without_cons != self.domain:
            yield without_cons

    def _reverse_subdomain(self):
        """97. Reverse subdomain."""
        if self.subdomain:
            yield '.'.join([self.subdomain[::-1], self.domain, self.tld])

    def _random_insertion(self):
        """98. Random character insertion."""
        chars = 'abcdefghijklmnopqrstuvwxyz'
        for _ in range(5):
            pos = random.randint(0, len(self.domain))
            c = random.choice(chars)
            yield self.domain[:pos] + c + self.domain[pos:]

    def _random_deletion(self):
        """99. Random deletion."""
        if len(self.domain) > 2:
            for _ in range(3):
                pos = random.randint(0, len(self.domain) - 1)
                yield self.domain[:pos] + self.domain[pos+1:]

    def _combination_attack(self):
        """100. Combine multiple techniques."""
        techniques = [
            self._homoglyph, self._cyrillic, self._greek,
            self._leetspeak, self._replacement, self._insertion
        ]
        for _ in range(20):
            selected = random.sample(techniques, 2)
            result = self.domain
            for tech in selected:
                try:
                    variants = list(tech())
                    if variants:
                        result = random.choice(variants)
                except:
                    continue
            if result != self.domain:
                yield result

    def _common_misspellings(self):
        """101. Common misspellings."""
        misspellings = {
            'en': ['an', 'in', 'on'],
            'er': ['re', 'ar', 'ir'],
            'es': ['se', 'as'],
            'le': ['el', 'al'],
            'al': ['el', 'il', 'ol'],
            'or': ['er', 'ar', 'ir'],
            'ai': ['ia', 'ae'],
            'ei': ['ie', 'ia'],
            'ou': ['uo', 'ow'],
            'ie': ['ei', 'e', 'y'],
            'tion': ['cion', 'sion', 'tiong'],
            'ing': ['eng', 'ang', 'ing'],
            'ment': ['mant', 'mento'],
            'ness': ['nes', 'nesse'],
            'ful': ['full', 'phul'],
            'ous': ['us', 'ious'],
            'able': ['ible', 'abel'],
            'ible': ['able', 'eble'],
            'ence': ['ance', 'ense'],
            'ance': ['ence', 'anse'],
            'tion': ['cion', 'sion'],
            'sion': ['tion', 'cion']
        }
        for i in range(len(self.domain) - 1):
            for j in range(2, min(5, len(self.domain) - i + 1)):
                substring = self.domain[i:i+j]
                for miss in misspellings.get(substring, []):
                    yield self.domain[:i] + miss + self.domain[i+j:]

    def _double_letter(self):
        """102. Double letters (common typo)."""
        for i, c in enumerate(self.domain):
            if c.isalpha():
                yield self.domain[:i] + c + c + self.domain[i+1:]

    def _missing_letter(self):
        """103. Missing letters (common typo)."""
        for i in range(len(self.domain)):
            yield self.domain[:i] + self.domain[i+1:]

    def _extra_letter(self):
        """104. Extra letter (common typo)."""
        chars = 'abcdefghijklmnopqrstuvwxyz'
        for i in range(len(self.domain) + 1):
            for c in chars:
                yield self.domain[:i] + c + self.domain[i:]

    def _letter_swap_adjacent(self):
        """105. Adjacent letter swap (common typo)."""
        for i in range(len(self.domain) - 1):
            if self.domain[i] != self.domain[i+1]:
                yield self.domain[:i] + self.domain[i+1] + self.domain[i] + self.domain[i+2:]

    def _letter_swap_distance2(self):
        """106. Swap letters at distance 2."""
        for i in range(len(self.domain) - 2):
            yield self.domain[:i] + self.domain[i+2] + self.domain[i+1] + self.domain[i] + self.domain[i+3:]

    # ========================================================================
    # GENERATE ALL PERMUTATIONS
    # ========================================================================

    def generate(self, fuzzers=None):
        """Generate all permutations - 100+ techniques."""
        self.domains = set()
        self.permutation_map = {}
        self.priority_weights = {}

        # Original domain
        original = '.'.join(filter(None, [self.subdomain, self.domain, self.tld]))
        self.domains.add(original)
        self.permutation_map[original] = 'original'
        self.priority_weights[original] = 1.0

        # All 100+ fuzzing techniques
        all_fuzzers = [
            'homoglyph', 'cyrillic', 'greek', 'arabic', 'hebrew',
            'leetspeak', 'uppercase-lowercase', 'accented-chars', 
            'double-characters', 'triple-characters',
            'insertion', 'omission', 'double-omission', 
            'insert-common', 'insert-number', 'insert-special',
            'delete-vowel', 'delete-consonant', 
            'insert-before', 'insert-after',
            'replacement', 'double-replacement', 'similar-looking',
            'number-for-letter', 'letter-for-number',
            'vowel-swap', 'consonant-swap',
            'adjacent-swap', 'word-boundary-swap', 'case-variation',
            'transposition', 'reverse', 'partial-reverse', 
            'shuffle', 'swap-near', 'swap-far',
            'rotate-left', 'rotate-right', 'mirror', 'upside-down',
            'hyphenation', 'double-hyphen', 'hyphen-substitution',
            'dot-substitution', 'missing-dots', 'double-dot',
            'subdomain-with-dot', 'subdomain-swap', 'tld-as-subdomain',
            'remove-subdomain',
            'tld-swap', 'tld-remove', 'tld-repeat', 'tld-without-dot',
            'tld-typos', 'tld-append', 'tld-prepend',
            'tld-misspelling', 'tld-combined', 'tld-duplicate',
            'prefix', 'suffix', 'prefix-suffix',
            'prefix-repeat', 'suffix-repeat',
            'prefix-typo', 'suffix-typo', 'prefix-suffix-swap',
            'numeric-prefix', 'numeric-suffix',
            'dictionary-insert', 'dictionary-replace',
            'brand-names', 'common-typos', 'word-boundary',
            'compound-words', 'word-split', 'word-merge',
            'keyword-insert', 'keyword-swap',
            'adding-numbers', 'number-replacement',
            'year-append', 'date-format',
            'special-characters', 'emoji-insert',
            'leet-advanced', 'symbol-substitution',
            'ascii-art-substitution', 'combining-diacritics',
            'bitsquatting', 'homoglyph-advanced',
            'domain-dots', 'domain-underscore',
            'domain-without-vowels', 'domain-without-consonants',
            'reverse-subdomain', 'random-insertion', 'random-deletion',
            'combination-attack', 'common-misspellings',
            'double-letter', 'missing-letter', 'extra-letter',
            'letter-swap-adjacent', 'letter-swap-distance2'
        ]
        
        fuzz_list = fuzzers if fuzzers else all_fuzzers

        for f_name in fuzz_list:
            try:
                f = getattr(self, '_' + f_name.replace('-', '_'))
                for domain in f():
                    if not self.tld:
                        full_domain = domain
                    else:
                        full_domain = '.'.join(filter(None, [self.subdomain, domain, self.tld]))

                    if validate_domain(full_domain) and full_domain not in self.domains:
                        self.domains.add(full_domain)
                        self.permutation_map[full_domain] = f_name
                        self.priority_weights[full_domain] = self._get_priority(f_name)
            except AttributeError:
                continue

        # Convert to punycode
        puny_domains = set()
        for domain in self.domains:
            try:
                if MODULES['idna']:
                    puny = idna.encode(domain).decode()
                else:
                    puny = domain
                if validate_domain(puny):
                    puny_domains.add(puny)
                    if puny != domain:
                        self.permutation_map[puny] = self.permutation_map.get(domain, 'punycode')
                        self.priority_weights[puny] = self.priority_weights.get(domain, 0.5)
            except:
                pass
        self.domains = puny_domains

    def _get_priority(self, technique):
        """Get priority weight for a fuzzing technique."""
        priority_map = {
            'homoglyph': 1.0, 'cyrillic': 0.95, 'greek': 0.90,
            'arabic': 0.88, 'hebrew': 0.87, 'leetspeak': 0.85,
            'similar-looking': 0.92, 'accented-chars': 0.89,
            'homoglyph-advanced': 0.93, 'combining-diacritics': 0.86,
            'bitsquatting': 0.80, 'common-typos': 0.78,
            'common-misspellings': 0.76, 'dictionary-insert': 0.84,
            'dictionary-replace': 0.82, 'brand-names': 0.83,
            'prefix': 0.75, 'suffix': 0.75, 'prefix-suffix': 0.77,
            'tld-swap': 0.79, 'tld-typos': 0.74, 'tld-misspelling': 0.72,
            'transposition': 0.70, 'replacement': 0.68,
            'insertion': 0.65, 'omission': 0.64,
            'addition': 0.63, 'repetition': 0.62,
            'combination-attack': 0.90, 'random-insertion': 0.60,
            'random-deletion': 0.58, 'shuffle': 0.55,
            'reverse': 0.50, 'mirror': 0.52, 'upside-down': 0.51,
            'case-variation': 0.45, 'vowel-swap': 0.44,
            'consonant-swap': 0.43, 'number-replacement': 0.42,
            'double-letter': 0.41, 'missing-letter': 0.40,
            'extra-letter': 0.40, 'letter-swap-adjacent': 0.48,
            'letter-swap-distance2': 0.47, 'hyphenation': 0.50,
            'dot-substitution': 0.46, 'domain-dots': 0.45,
            'symbol-substitution': 0.38, 'ascii-art-substitution': 0.35,
            'emoji-insert': 0.30, 'year-append': 0.28,
            'date-format': 0.25, 'special-characters': 0.22,
            'numeric-prefix': 0.33, 'numeric-suffix': 0.32,
            'domain-without-vowels': 0.37, 'domain-without-consonants': 0.36,
            'uppercase-lowercase': 0.44, 'double-characters': 0.43,
            'triple-characters': 0.42, 'double-omission': 0.41,
            'insert-common': 0.40, 'insert-number': 0.39,
            'insert-special': 0.38, 'delete-vowel': 0.37,
            'delete-consonant': 0.36, 'insert-before': 0.35,
            'insert-after': 0.34, 'double-replacement': 0.33,
            'number-for-letter': 0.32, 'letter-for-number': 0.31,
            'adjacent-swap': 0.30, 'word-boundary-swap': 0.29,
            'partial-reverse': 0.28, 'swap-near': 0.27,
            'swap-far': 0.26, 'rotate-left': 0.25,
            'rotate-right': 0.24, 'double-hyphen': 0.23,
            'hyphen-substitution': 0.22, 'double-dot': 0.21,
            'subdomain-with-dot': 0.20, 'subdomain-swap': 0.19,
            'tld-as-subdomain': 0.18, 'remove-subdomain': 0.17,
            'tld-remove': 0.16, 'tld-repeat': 0.15,
            'tld-without-dot': 0.14, 'tld-append': 0.13,
            'tld-prepend': 0.12, 'tld-combined': 0.11,
            'tld-duplicate': 0.10, 'prefix-repeat': 0.09,
            'suffix-repeat': 0.08, 'prefix-typo': 0.07,
            'suffix-typo': 0.06, 'prefix-suffix-swap': 0.05,
            'word-boundary': 0.04, 'compound-words': 0.03,
            'word-split': 0.02, 'word-merge': 0.01,
            'keyword-insert': 0.00, 'keyword-swap': 0.00,
            'adding-numbers': 0.00, 'reverse-subdomain': 0.00
        }
        return priority_map.get(technique, 0.5)

    def permutations(self, registered=False, unicode=False):
        """Get permutations list."""
        result = []
        for domain in sorted(self.domains):
            if MODULES['idna'] and unicode:
                display = idna.decode(domain)
            else:
                display = domain
            entry = {
                'domain': display,
                'fuzzer': self.permutation_map.get(domain, 'unknown'),
                'punycode': domain,
                'priority': self.priority_weights.get(domain, 0.5),
                'registered': registered
            }
            result.append(entry)
        return result

# ============================================================================
# COMPLETE URL PARSER CLASS
# ============================================================================

class UrlParser:
    """Complete URL parser class."""
    
    def __init__(self, url):
        if not url:
            raise TypeError('argument has to be non-empty string')
        u = urllib.parse.urlparse(url if '://' in url else '//' + url, scheme='http')
        self.scheme = u.scheme.lower()
        if self.scheme not in ('http', 'https'):
            raise ValueError('invalid scheme')
        self.domain = u.hostname.lower()
        try:
            self.domain = idna.encode(self.domain).decode()
        except:
            pass
        if not validate_domain(self.domain):
            raise ValueError('invalid domain name')
        self.username = u.username
        self.password = u.password
        self.port = u.port
        self.path = u.path
        self.query = u.query
        self.fragment = u.fragment

    def full_uri(self, domain=None):
        """Return full URI."""
        uri = '{}://'.format(self.scheme)
        if self.username:
            uri += self.username
            if self.password:
                uri += ':{}'.format(self.password)
            uri += '@'
        uri += self.domain if not domain else domain
        if self.port:
            uri += ':{}'.format(self.port)
        if self.path:
            uri += self.path
        if self.query:
            uri += '?{}'.format(self.query)
        if self.fragment:
            uri += '#{}'.format(self.fragment)
        return uri

# ============================================================================
# PRINT FUNCTIONS
# ============================================================================

def print_banner():
    """Print the complete banner with ASCII art."""
    modules_loaded = sum(1 for v in MODULES.values() if v)
    
    banner = f"""
{Colors.RED}
     ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣀⣤⣤⣤⣶⣶⣶⣾⣿⣿⣿⣿⣿⣶⣶⣶⣶⣶⣤⣤⣄⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣦⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣴⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⣠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣄⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⡀⠀⠀⠀⠀⠀
⠀⠀⠀⣠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⡀⠀⠀⠀
⠀⠀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⡀⠀⠀
⠀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡀⠀
⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⠀
⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡄
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⡟⠀⠀⠉⠛⠻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠛⠉⠈⠁⠉⣿⣿⣿⣿
⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠉⠛⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⠉⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⡟
⢸⣿⣿⣿⣇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠛⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣿⣿⣿⡇
⠀⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠙⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣼⣿⣿⣿⠀
⠀⢹⣿⣿⣿⣷⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠛⠻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⠉⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣰⣿⣿⣿⠇⠀
⠀⠀⢻⣿⣿⣿⣿⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠉⠛⣿⣿⣿⣿⡿⠛⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⣾⣿⣿⣿⣿⠀⠀
⠀⠀⠀⢻⣿⣿⣿⣿⣿⣷⣶⣦⣤⣄⣀⣀⣀⠀⠀⠀⠀⠀⢀⣀⣤⣴⣶⣿⣿⠟⡿⣿⣿⣷⣶⣤⣄⡀⠀⠀⠀⠀⠀⣀⣀⣀⣤⣤⣤⣴⣶⣿⣿⣿⣿⣿⣿⡏⠀⠀
⠀⠀⠀⠀⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠁⠀⡇⠈⢹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠁⠀⠀
⠀⠀⠀⠀⢀⣬⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠃⠀⠀⡇⠀⠀⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡏⠀⠀⠀
⠀⠀⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠁⠀⠀⣼⣧⡀⠀⠀⠙⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⡀⠀⠀
⠀⠀⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⠻⢿⣿⣿⣿⣿⣿⣧⠀⢀⣼⣿⣿⣷⡄⠀⣠⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠿⠿⠿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀
⠀⠀⠀⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠋⠁⠀⠀⠀⣸⣿⣿⣿⣿⣿⣿⣧⣀⣿⣿⣿⣿⣧⣰⣿⣿⣿⣿⣿⣿⣿⣿⡟⠁⠀⠀⠀⠀⠀⠈⢻⣿⣿⣿⣿⣿⡟⠀⠀⠀
⠀⠀⠀⠘⣿⣿⣿⣿⣿⣿⣿⠟⠁⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⡿⠋⠀⠀⠀⠀
⠀⠀⠀⠀⠈⠻⣿⠟⠋⠉⠛⠷⠶⠒⠀⠀⠀⠀⠀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠀⠀⠀⠀⠀⠀⠈⠻⠿⠛⠉⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢹⣿⣿⣿⣿⡟⢻⣿⣿⣿⣿⠛⢻⣿⣿⡏⠀⢀⣽⣿⣿⣿⣿⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⣿⣿⣿⣿⡇⢸⣿⣿⣿⣿⠀⣿⣿⣿⡇⠀⣿⣿⣿⣿⣿⠏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⢸⣿⣿⣿⣿⠀⣿⣿⣿⣇⠀⣿⣿⣿⣿⡿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⣿⢸⣿⣿⣿⣿⠀⣿⣿⣿⣿⢰⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢹⣿⣿⣿⣿⢸⣿⣿⣿⣿⠀⣿⣿⣿⣿⢸⣿⣿⣿⣿⠇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⢸⣿⣿⣿⣿⠀⣿⣿⣿⣿⢸⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⣿⣿⣿⣿⠀⣿⣿⣿⣿⠀⣿⣿⣿⡿⢸⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⣿⣿⠀⣿⣿⣿⣿⠀⣿⣿⣿⡇⢸⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢿⣿⣿⣿⠀⣿⣿⣿⣿⠀⣿⣿⣿⡇⢸⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⠀⣿⣿⣿⡿⠀⣿⣿⣿⡇⢸⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⠀⣿⣿⣿⡇⠀⣿⣿⣿⠇⢸⣿⣿⣿⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠸⣿⣿⣿⠀⢻⣿⣿⡇⠀⣿⣿⣿⠀⢸⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠻⠋⠀⠸⣿⣿⠇⠀⢿⣿⠏⠀⠸⠿⠿⠿⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀                                                                             
 _   _ ______ _________          ___    _ _____  _____ _____  ______ _____
| \ | |  ____|__   __\ \        / / |  | |_   _|/ ____|  __ \|  ____|  __ \\
|  \| | |__     | |   \ \  /\  / /| |__| | | | | (___ | |__) | |__  | |__) |
| . ` |  __|    | |    \ \/  \/ / |  __  | | |  \___ \|  ___/|  __| |  _  /
| |\  | |____   | |     \  /\  /  | |  | |_| |_ ____) | |    | |____| | \ \\
|_| \_|______|  |_|      \/  \/   |_|  |_|_____|_____/|_|    |______|_|  \_\\
                  Advanced Phishing Domain Scanner                            
                     Threat Intelligence Platform                             
                         ULTIMATE EDITION                                    
                     100+ PERMUTATION TECHNIQUES

{Colors.END}

{Colors.CYAN}[SYSTEM] NetWhisper... initialized
[SYSTEM] Author: {AUTHOR}
[SYSTEM] 100+ Domain Variation Techniques: LOADED
[SYSTEM] Threat Intelligence Engine: ACTIVE
[SYSTEM] All-in-One Mode: ENABLED
[SYSTEM] Complete Feature Set: LOADED
[SYSTEM] DNS: {'dnspython (fast)' if MODULES['dnspython'] else 'socket (fallback)'}
[SYSTEM] IDN: {'enabled' if MODULES['idna'] else 'disabled'}
[SYSTEM] Modules Loaded: {modules_loaded}/{len(MODULES)}
[SYSTEM] Fuzzing Techniques: 106{Colors.END}
"""
    print(banner)

def print_summary(summary, target):
    """Print scan summary."""
    print(f"""
{Colors.BOLD}╔══════════════════════════════════════════════════════════════════════════════════╗
║                              SCAN COMPLETE                                                 ║
╠══════════════════════════════════════════════════════════════════════════════════╣{Colors.END}
{Colors.CYAN}  Target Domain     : {target}
  Total Permutations: {summary['total']}
  Registered Domains: {summary['registered']}
  Scan Duration     : {format_duration(summary['duration'])}{Colors.END}
{Colors.BOLD}╠══════════════════════════════════════════════════════════════════════════════════╣
║  THREAT LEVEL SUMMARY                                                                     ║
╠══════════════════════════════════════════════════════════════════════════════════╣{Colors.END}
{Colors.RED}  🔴 CRITICAL        : {summary['critical']}{Colors.END}
{Colors.YELLOW}  🟠 HIGH           : {summary['high']}{Colors.END}
{Colors.YELLOW}  🟡 MEDIUM         : {summary['medium']}{Colors.END}
{Colors.CYAN}  🔵 LOW            : {summary['low']}{Colors.END}
{Colors.GREEN}  🟢 SAFE           : {summary['safe']}{Colors.END}
{Colors.BOLD}╚══════════════════════════════════════════════════════════════════════════════════╝{Colors.END}
""")

def print_results_table(results, show_all=False, max_display=100):
    """Print results table."""
    if not results:
        return
    
    filtered = []
    for domain, data in results.items():
        if show_all or data.get('dns_a') or data.get('dns_aaaa'):
            filtered.append((domain, data))
    
    if not filtered:
        print(f"\n{Colors.YELLOW}⚠ No registered domains found.{Colors.END}")
        return
    
    filtered.sort(key=lambda x: x[1].get('risk_score', 0), reverse=True)
    
    print(f"\n{Colors.BOLD}{'=' * 130}{Colors.END}")
    print(f"{Colors.BOLD}DOMAIN{' ' * 45}IP{' ' * 18}RISK   LEVEL{' ' * 8}FUZZER{' ' * 12}INFO{Colors.END}")
    print(f"{Colors.BOLD}{'-' * 130}{Colors.END}")
    
    display_count = min(len(filtered), max_display)
    
    for domain, data in filtered[:display_count]:
        ip = data.get('dns_a', [''])[0] if data.get('dns_a') else ''
        risk = data.get('risk_score', 0)
        level = data.get('threat_level', 'UNKNOWN')
        fuzzer = data.get('fuzzer', '')
        
        level_colors = {
            'CRITICAL': Colors.RED,
            'HIGH': Colors.YELLOW,
            'MEDIUM': Colors.YELLOW,
            'LOW': Colors.CYAN,
            'SAFE': Colors.GREEN
        }
        color = level_colors.get(level, Colors.END)
        
        domain_display = domain[:45] + '...' if len(domain) > 45 else domain.ljust(48)
        ip_display = ip[:18] + '...' if len(ip) > 18 else ip.ljust(20)
        fuzzer_display = fuzzer[:15] if fuzzer else ''
        
        info_parts = []
        if data.get('geoip'):
            info_parts.append(f"📍{data['geoip']}")
        if data.get('mx_spy'):
            info_parts.append("🕵️")
        if data.get('phash'):
            info_parts.append(f"🎨{data['phash']}")
        if data.get('ssdeep'):
            info_parts.append(f"🔍{data['ssdeep']}")
        if data.get('ssl_valid'):
            info_parts.append("🔒")
        info_display = ' '.join(info_parts)[:25]
        
        print(f"{color}{domain_display}{ip_display}{risk:>3}%   {level:<8}{Colors.END} {fuzzer_display:<15} {info_display}")
    
    if len(filtered) > max_display:
        print(f"{Colors.GRAY}... and {len(filtered) - max_display} more domains{Colors.END}")
    
    print(f"{Colors.BOLD}{'=' * 130}{Colors.END}")

def export_results(results, target, filename, format_type='json'):
    """Export results to file."""
    data = {
        'target': target,
        'timestamp': datetime.now().isoformat(),
        'author': AUTHOR,
        'version': VERSION,
        'total': len(results),
        'registered': sum(1 for r in results.values() if r.get('dns_a') or r.get('dns_aaaa')),
        'results': []
    }
    
    for domain, r in results.items():
        data['results'].append({
            'domain': domain,
            'fuzzer': r.get('fuzzer', ''),
            'ip': r.get('dns_a', [''])[0] if r.get('dns_a') else '',
            'dns_a': r.get('dns_a', []),
            'dns_aaaa': r.get('dns_aaaa', []),
            'dns_ns': r.get('dns_ns', []),
            'dns_mx': r.get('dns_mx', []),
            'geoip': r.get('geoip', ''),
            'banner_http': r.get('banner_http', ''),
            'banner_smtp': r.get('banner_smtp', ''),
            'mx_spy': r.get('mx_spy', False),
            'phash': r.get('phash', ''),
            'ssdeep': r.get('ssdeep', ''),
            'tlsh': r.get('tlsh', ''),
            'whois_created': r.get('whois_created', ''),
            'whois_registrar': r.get('whois_registrar', ''),
            'whois_expires': r.get('whois_expires', ''),
            'whois_nameservers': r.get('whois_nameservers', ''),
            'whois_status': r.get('whois_status', ''),
            'risk_score': r.get('risk_score', 0),
            'threat_level': r.get('threat_level', 'SAFE'),
            'ssl_valid': r.get('ssl_valid', False)
        })
    
    try:
        if format_type == 'json':
            with open(filename, 'w') as f:
                json.dump(data, f, indent=2)
        elif format_type == 'csv':
            import csv
            with open(filename, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Domain', 'IP', 'Risk', 'Level', 'Fuzzer', 'GeoIP', 
                               'Registrar', 'Created', 'Expires', 'MX Spy', 'pHash', 'SSDeep', 'SSL'])
                for domain, r in results.items():
                    writer.writerow([
                        domain,
                        r.get('dns_a', [''])[0] if r.get('dns_a') else '',
                        r.get('risk_score', 0),
                        r.get('threat_level', ''),
                        r.get('fuzzer', ''),
                        r.get('geoip', ''),
                        r.get('whois_registrar', ''),
                        r.get('whois_created', ''),
                        r.get('whois_expires', ''),
                        r.get('mx_spy', False),
                        r.get('phash', ''),
                        r.get('ssdeep', ''),
                        r.get('ssl_valid', False)
                    ])
        elif format_type == 'txt':
            with open(filename, 'w') as f:
                f.write(f"NetWhisper ULTIMATE Scan Report\n")
                f.write(f"Target: {target}\n")
                f.write(f"Generated: {datetime.now().isoformat()}\n")
                f.write(f"Author: {AUTHOR}\n")
                f.write(f"Version: {VERSION}\n")
                f.write(f"{'='*80}\n\n")
                for domain, r in results.items():
                    if r.get('dns_a') or r.get('dns_aaaa'):
                        f.write(f"{domain} | {r.get('dns_a', [''])[0]} | {r.get('risk_score', 0)}% | {r.get('threat_level', '')}\n")
        
        print(f"\n{Colors.GREEN}✅ Results exported to: {filename}{Colors.END}")
    except Exception as e:
        print(f"\n{Colors.RED}❌ Export failed: {e}{Colors.END}")

# ============================================================================
# SESSION MANAGER - Complete
# ============================================================================

class Session:
    """Complete scan session manager."""
    
    def __init__(self, url, options=None):
        self.id = str(uuid4())
        self.timestamp = int(time.time())
        self.url = UrlParser(url)
        self.options = options or {}
        self.jobs = queue.Queue()
        self.results = {}
        self.threads = []
        self.complete = 0
        self.total = 0
        self.running = True
        self.whois_done = False
        self.start_time = time.time()

        # Generate permutations
        self.fuzzer = Fuzzer(self.url.domain, dictionary=DICTIONARY, tld_dictionary=ABUSED_TLDS)
        self.fuzzer.generate()
        self.permutations = self.fuzzer.permutations(unicode=True)

        # Initialize LSH if requested
        self.lsh_init = ''
        self.lsh_effective_url = ''
        if self.options.get('lsh'):
            try:
                url_to_fetch = self.options.get('lsh_url', self.url.full_uri())
                r = UrlOpener(url_to_fetch, timeout=5.0)
                if self.options['lsh'] == 'ssdeep' and MODULES.get('ssdeep', False):
                    self.lsh_init = ssdeep.hash(r.normalized_content)
                elif self.options['lsh'] == 'tlsh' and MODULES.get('tlsh', False):
                    self.lsh_init = tlsh.hash(r.normalized_content)
                self.lsh_effective_url = r.url.split('?')[0]
                if self.lsh_init in (None, '', 'TNULL', '3::'):
                    self.options['lsh'] = None
            except:
                self.options['lsh'] = None

        # Initialize pHash if requested
        self.phash_init = None
        if self.options.get('phash'):
            try:
                url_to_fetch = self.options.get('phash_url', self.url.full_uri())
                browser = HeadlessBrowser(useragent=USER_AGENT_STRING)
                browser.get(url_to_fetch)
                screenshot = browser.screenshot()
                self.phash_init = pHash(BytesIO(screenshot))
                browser.stop()
            except:
                self.options['phash'] = None

    def scan(self):
        """Start scanning."""
        for item in self.permutations:
            self.jobs.put(item)

        self.total = self.jobs.qsize()
        thread_count = min(16, max(1, self.total // 10))

        for _ in range(thread_count):
            worker_options = self.options.copy()
            worker_options['lsh_init'] = self.lsh_init
            worker_options['lsh_url'] = self.lsh_effective_url
            
            worker = Scanner(self.jobs, self.results, self.url, worker_options)
            worker.start()
            self.threads.append(worker)

        def monitor():
            while self.running:
                time.sleep(0.2)
                self.complete = self.total - self.jobs.qsize()
                if self.jobs.empty():
                    break
                alive = any(t.is_alive() for t in self.threads)
                if not alive and not self.jobs.empty():
                    break
            self.complete = self.total

        t = threading.Thread(target=monitor)
        t.daemon = True
        t.start()

    def stop(self):
        """Stop scanning."""
        self.running = False
        for t in self.threads:
            t.stop()
        while not self.jobs.empty():
            try:
                self.jobs.get_nowait()
            except queue.Empty:
                break

    def get_summary(self):
        """Get scan summary."""
        total = len(self.results)
        registered = sum(1 for r in self.results.values() if r.get('dns_a') or r.get('dns_aaaa'))
        critical = sum(1 for r in self.results.values() if r.get('threat_level') == 'CRITICAL')
        high = sum(1 for r in self.results.values() if r.get('threat_level') == 'HIGH')
        medium = sum(1 for r in self.results.values() if r.get('threat_level') == 'MEDIUM')
        low = sum(1 for r in self.results.values() if r.get('threat_level') == 'LOW')
        safe = sum(1 for r in self.results.values() if r.get('threat_level') == 'SAFE')
        
        return {
            'total': total,
            'registered': registered,
            'critical': critical,
            'high': high,
            'medium': medium,
            'low': low,
            'safe': safe,
            'duration': time.time() - self.start_time
        }

    def domains(self):
        """Get domain results."""
        result = []
        for item in self.permutations:
            domain = item['domain']
            entry = {
                'domain': domain,
                'fuzzer': item.get('fuzzer', ''),
                'punycode': item.get('punycode', domain),
                'priority': item.get('priority', 0.5)
            }
            
            if domain in self.results:
                r = self.results[domain]
                for key in ['dns_a', 'dns_aaaa', 'dns_ns', 'dns_mx', 'geoip', 
                           'banner_http', 'banner_smtp', 'mx_spy', 'phash', 
                           'ssdeep', 'tlsh', 'whois_created', 'whois_registrar',
                           'whois_expires', 'whois_nameservers', 'whois_status',
                           'risk_score', 'threat_level', 'ssl_valid']:
                    if key in r:
                        entry[key] = r[key]
            
            result.append(entry)
        return result

# ============================================================================
# SCANNER CLASS
# ============================================================================

class Scanner(threading.Thread):
    """Complete scanner thread with all analysis features."""
    
    def __init__(self, queue, results, url, options=None):
        threading.Thread.__init__(self)
        self.daemon = True
        self.jobs = queue
        self.results = results
        self.url = url
        self.options = options or {}
        self._stop_event = threading.Event()
        self.id = int.from_bytes(os.urandom(4), 'little')
        self.resolv = None
        self.geo = None
        self.browser = None
        self.whois = Whois()
        self.lsh_init = ''
        self.lsh_effective_url = ''
        self.phash_init = None
        self.screenshot_dir = None

        # Initialize DNS resolver with error handling for Termux
        if MODULES['dnspython']:
            try:
                self.resolv = Resolver()
                self.resolv.search = []
                self.resolv.lifetime = 5.0
                self.resolv.timeout = 2.5
                if self.options.get('nameservers'):
                    self.resolv.nameservers = self.options['nameservers']
                else:
                    # Try common DNS servers if resolv.conf not found
                    try:
                        self.resolv.nameservers = ['8.8.8.8', '1.1.1.1', '9.9.9.9']
                    except:
                        pass
            except Exception as e:
                # Fallback to socket-based DNS
                self.resolv = None
                MODULES['dnspython'] = False

        # Initialize GeoIP
        if MODULES.get('geoip2', False) and self.options.get('geoip', True):
            try:
                self.geo = geoip2.database.Reader('GeoLite2-Country.mmdb')
            except:
                try:
                    import GeoIP
                    self.geo = GeoIP.new(GeoIP.GEOIP_MEMORY_CACHE)
                except:
                    self.geo = None

        # Initialize browser
        if MODULES.get('selenium', False) and MODULES.get('pil', False) and self.options.get('phash', False):
            try:
                self.browser = HeadlessBrowser(useragent=USER_AGENT_STRING)
            except:
                self.browser = None

        # Initialize LSH
        if self.options.get('lsh'):
            self.lsh_init = self.options.get('lsh_init', '')
            self.lsh_effective_url = self.options.get('lsh_url', '')

    def stop(self):
        self._stop_event.set()

    def is_stopped(self):
        return self._stop_event.is_set()

    def _send_recv_tcp(self, host, port, data=b'', timeout=2.0, recv_bytes=1024):
        """TCP send/receive utility."""
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        resp = b''
        try:
            sock.connect((host, port))
            if data:
                sock.send(data)
            resp = sock.recv(recv_bytes)
        except:
            pass
        finally:
            sock.close()
        return resp.decode('utf-8', errors='ignore')

    def _banner_http(self, ip, vhost):
        """Get HTTP banner."""
        response = self._send_recv_tcp(ip, 80,
            f'HEAD / HTTP/1.1\r\nHost: {vhost}\r\nUser-Agent: {USER_AGENT_STRING}\r\n\r\n'.encode())
        if not response:
            return ''
        for line in response.splitlines():
            if line.lower().startswith('server: '):
                return line[8:]
        return ''

    def _banner_smtp(self, mx):
        """Get SMTP banner."""
        response = self._send_recv_tcp(mx, 25)
        if not response:
            return ''
        hello = response.splitlines()[0]
        if hello.startswith('220'):
            return hello[4:].strip()
        return ''

    def _mxcheck(self, mxhost, domain_from, domain_rcpt):
        """Check if MX can intercept emails."""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5.0)
            sock.connect((mxhost, 25))
        except:
            return False

        for cmd in [
            f'EHLO {mxhost}\r\n',
            f'MAIL FROM: test@{domain_from}\r\n',
            f'RCPT TO: test@{domain_rcpt}\r\n',
        ]:
            try:
                resp = sock.recv(512)
                if not resp or resp[0] != 0x32:
                    break
                sock.send(cmd.encode())
            except:
                break
        else:
            sock.close()
            return True
        sock.close()
        return False

    def _check_ssl(self, domain, ip):
        """Check SSL/TLS certificate."""
        result = {'valid': False, 'issuer': '', 'subject': '', 'expires': '', 'san': []}
        try:
            import ssl
            ctx = ssl.create_default_context()
            with socket.create_connection((ip, 443), timeout=5) as sock:
                with ctx.wrap_socket(sock, server_hostname=domain) as ssock:
                    cert = ssock.getpeercert()
                    if cert:
                        result['valid'] = True
                        result['issuer'] = dict(x[0] for x in cert.get('issuer', []))
                        result['subject'] = dict(x[0] for x in cert.get('subject', []))
                        result['expires'] = cert.get('notAfter', '')
                        result['san'] = cert.get('subjectAltName', [])
        except:
            pass
        return result

    def run(self):
        """Main scanner loop - Complete."""
        while not self.is_stopped():
            try:
                task = self.jobs.get(block=False)
            except queue.Empty:
                break

            domain = task['domain']
            result = {
                'domain': domain,
                'fuzzer': task['fuzzer'],
                'priority': task.get('priority', 0.5)
            }

            # DNS Resolution - Try dnspython first, fallback to socket
            if MODULES['dnspython'] and self.resolv:
                nxdomain = False
                try:
                    ns = self.resolv.resolve(domain, rdtype=dns.rdatatype.NS)
                    result['dns_ns'] = answer_to_list(ns)
                except NXDOMAIN:
                    nxdomain = True
                except NoNameservers:
                    result['dns_ns'] = ['!ServFail']
                except DNSException:
                    pass

                if not nxdomain:
                    try:
                        a = self.resolv.resolve(domain, rdtype=dns.rdatatype.A)
                        result['dns_a'] = answer_to_list(a)
                    except NoNameservers:
                        result['dns_a'] = ['!ServFail']
                    except:
                        pass

                    try:
                        aaaa = self.resolv.resolve(domain, rdtype=dns.rdatatype.AAAA)
                        result['dns_aaaa'] = answer_to_list(aaaa)
                    except:
                        pass

                    try:
                        mx = self.resolv.resolve(domain, rdtype=dns.rdatatype.MX)
                        result['dns_mx'] = answer_to_list(mx)
                    except:
                        pass
            else:
                # Fallback to socket-based DNS resolution
                try:
                    ips = []
                    addrinfo = socket.getaddrinfo(domain, None, proto=socket.IPPROTO_TCP)
                    for _, _, _, _, sa in addrinfo:
                        ip = sa[0]
                        if '.' in ip and ip not in ips:
                            ips.append(ip)
                    if ips:
                        result['dns_a'] = ips[:5]
                except:
                    pass

            # Skip if not registered
            if not result.get('dns_a') and not result.get('dns_aaaa'):
                self.jobs.task_done()
                self.results[domain] = result
                continue

            # GeoIP
            if self.geo and result.get('dns_a'):
                try:
                    if hasattr(self.geo, 'country_by_addr'):
                        country = self.geo.country_by_addr(result['dns_a'][0])
                    else:
                        country = self.geo.country(result['dns_a'][0]).country.name
                    if country:
                        result['geoip'] = country.split(',')[0]
                except:
                    pass

            # SSL Certificate Check
            if result.get('dns_a'):
                ssl_info = self._check_ssl(domain, result['dns_a'][0])
                result['ssl_info'] = ssl_info
                result['ssl_valid'] = ssl_info.get('valid', False)

            # MX Spy Check
            if self.options.get('mxcheck', True) and result.get('dns_mx'):
                if domain != self.url.domain:
                    if self._mxcheck(result['dns_mx'][0], self.url.domain, domain):
                        result['mx_spy'] = True

            # HTTP Banners
            if self.options.get('banners', True) and result.get('dns_a'):
                banner = self._banner_http(result['dns_a'][0], domain)
                if banner:
                    result['banner_http'] = banner

            # SMTP Banners
            if self.options.get('banners', True) and result.get('dns_mx'):
                banner = self._banner_smtp(result['dns_mx'][0])
                if banner:
                    result['banner_smtp'] = banner

            # WHOIS
            if self.options.get('whois', True):
                try:
                    wreply = self.whois.whois(domain)
                    if wreply.get('creation_date'):
                        result['whois_created'] = wreply['creation_date'].strftime('%Y-%m-%d')
                    if wreply.get('registrar'):
                        result['whois_registrar'] = wreply['registrar']
                    if wreply.get('expiration_date'):
                        result['whois_expires'] = wreply['expiration_date'].strftime('%Y-%m-%d')
                    if wreply.get('name_servers'):
                        result['whois_nameservers'] = wreply['name_servers']
                    if wreply.get('status'):
                        result['whois_status'] = wreply['status']
                except:
                    pass

            # Risk Score - Complete
            risk = 0
            if result.get('dns_a'):
                risk += 30
                
                # Domain age
                if result.get('whois_created'):
                    try:
                        created = datetime.strptime(result['whois_created'], '%Y-%m-%d')
                        age = (datetime.now() - created).days
                        if age < 30:
                            risk += 40
                        elif age < 90:
                            risk += 20
                        elif age < 365:
                            risk += 10
                    except:
                        pass
                
                # Similarity
                sim = calculate_similarity(self.url.domain, domain)
                if sim > 0.85:
                    risk += 25
                elif sim > 0.70:
                    risk += 15
                elif sim > 0.50:
                    risk += 5
                
                # MX Spy
                if result.get('mx_spy'):
                    risk += 15
                
                # IDN
                if is_idn(domain):
                    risk += 10
                
                # SSL
                if result.get('ssl_valid'):
                    risk += 5

            result['risk_score'] = min(risk, 100)
            
            if result['risk_score'] >= 75:
                result['threat_level'] = 'CRITICAL'
            elif result['risk_score'] >= 55:
                result['threat_level'] = 'HIGH'
            elif result['risk_score'] >= 35:
                result['threat_level'] = 'MEDIUM'
            elif result['risk_score'] >= 15:
                result['threat_level'] = 'LOW'
            else:
                result['threat_level'] = 'SAFE'

            self.results[domain] = result
            self.jobs.task_done()

# ============================================================================
# HELPER CLASSES FOR SESSION
# ============================================================================

class UrlOpener:
    """Complete URL opener with content normalization."""
    
    def __init__(self, url, timeout=5.0, headers=None, verify=True):
        http_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9',
            'accept-encoding': 'gzip,identity',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
            'User-Agent': USER_AGENT_STRING
        }
        if headers:
            for h, v in headers.items():
                if h.lower() != 'accept-encoding':
                    http_headers[h.lower()] = v
        
        if verify:
            ctx = ssl.create_default_context()
        else:
            ctx = ssl._create_unverified_context()
        
        request = urllib.request.Request(url, headers=http_headers)
        with urllib.request.urlopen(request, timeout=timeout, context=ctx) as r:
            self.headers = r.headers
            self.code = r.code
            self.reason = r.reason
            self.url = r.url
            self.content = r.read()
        
        if self.content[:3] == b'\x1f\x8b\x08':
            self.content = gzip.decompress(self.content)
        
        self.normalized_content = self._normalize()

    def _normalize(self):
        """Normalize content for similarity comparison."""
        content = b' '.join(self.content.split())
        content = re.sub(b'(action|src|href)="[^"]*"', lambda m: m.group(0).split(b'=')[0] + b'=""', content)
        # Fixed escape sequence - using double backslash for regex
        content = re.sub(b'url\\([^)]*\\)', b'url()', content)
        return content

class pHash:
    """Perceptual hash for image similarity."""
    
    def __init__(self, image_data, hsize=8):
        if not MODULES.get('pil', False):
            raise ImportError('PIL module required for pHash')
        img = Image.open(BytesIO(image_data))
        img = img.convert('L').resize((hsize, hsize), Image.LANCZOS)
        pixels = list(img.getdata())
        avg = sum(pixels) / len(pixels)
        self.hash = ''.join('1' if p > avg else '0' for p in pixels)
        self.size = hsize

    def __sub__(self, other):
        """Hamming distance."""
        if self.size != other.size:
            return 100
        ham = sum(x != y for x, y in zip(self.hash, other.hash))
        e = 2.718281828459045
        result = int((1 + e**((self.size**2 - ham) / self.size**2) - e) * 100)
        return result if result > 0 else 0

    def __repr__(self):
        return '{:x}'.format(int(self.hash, base=2))

    def __int__(self):
        return int(self.hash, base=2)

class HeadlessBrowser:
    """Complete headless browser for screenshots."""
    
    WEBDRIVER_ARGUMENTS = (
        '--disable-dev-shm-usage',
        '--ignore-certificate-errors',
        '--headless',
        '--incognito',
        '--no-sandbox',
        '--disable-gpu',
        '--disable-extensions',
        '--disk-cache-size=0',
        '--aggressive-cache-discard',
        '--disable-notifications',
        '--disable-remote-fonts',
        '--disable-sync',
        '--window-size=1366,768',
        '--hide-scrollbars',
        '--disable-audio-output',
        '--dns-prefetch-disable',
        '--no-default-browser-check',
        '--disable-background-timer-throttling',
        '--disable-backgrounding-occluded-windows',
        '--disable-breakpad',
        '--disable-client-side-phishing-detection',
        '--disable-component-extensions-with-background-pages',
        '--disable-default-apps',
        '--disable-features=TranslateUI',
        '--disable-hang-monitor',
        '--disable-ipc-flooding-protection',
        '--disable-prompt-on-repost',
        '--disable-renderer-backgrounding',
        '--force-color-profile=srgb',
        '--metrics-recording-only',
        '--no-first-run',
        '--password-store=basic',
        '--use-mock-keychain',
        '--disable-blink-features=AutomationControlled',
    )
    
    def __init__(self, useragent=None):
        if not MODULES.get('selenium', False):
            raise ImportError('Selenium module required for HeadlessBrowser')
        
        chrome_options = Options()
        for opt in self.WEBDRIVER_ARGUMENTS:
            chrome_options.add_argument(opt)
        
        proxies = urllib.request.getproxies()
        if proxies:
            proxy_string = ';'.join(['{}={}'.format(scheme, url) for scheme, url in proxies.items()])
            chrome_options.add_argument('--proxy-server={}'.format(proxy_string))
        
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.set_page_load_timeout(12.0)
        
        if useragent:
            self.driver.execute_cdp_cmd('Network.setUserAgentOverride', {'userAgent': useragent})
        
        self.get = self.driver.get
        self.screenshot = self.driver.get_screenshot_as_png

    def stop(self):
        """Stop the browser."""
        try:
            self.driver.close()
            self.driver.quit()
        except:
            pass
        try:
            pid = True
            while pid:
                pid, status = os.waitpid(-1, os.WNOHANG)
        except ChildProcessError:
            pass

    def __del__(self):
        self.stop()

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def signal_handler(sig, frame):
    """Handle Ctrl+C."""
    print(f"\n\n{Colors.YELLOW}[!] Scan interrupted by user{Colors.END}")
    sys.exit(0)

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='NetWhisper Ultimate - Advanced Phishing Domain Scanner with 100+ Permutation Techniques',
        epilog=f'Author: {AUTHOR} | Version: {VERSION}'
    )
    parser.add_argument('-d', '--domain', required=True, help='Domain to scan (e.g., example.com)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-f', '--format', choices=['json', 'csv', 'txt'], default='json',
                       help='Output format (default: json)')
    parser.add_argument('-t', '--threads', type=int, default=16, help='Number of threads (default: 16)')
    parser.add_argument('--all', action='store_true', help='Show all domains (including unregistered)')
    parser.add_argument('--quiet', action='store_true', help='Quiet mode (minimal output)')
    parser.add_argument('--no-geoip', action='store_true', help='Disable GeoIP')
    parser.add_argument('--no-banner', action='store_true', help='Disable banners')
    parser.add_argument('--no-mxcheck', action='store_true', help='Disable MX spy')
    parser.add_argument('--no-whois', action='store_true', help='Disable WHOIS')
    parser.add_argument('--phash', action='store_true', help='Enable pHash visual similarity')
    parser.add_argument('--lsh', choices=['ssdeep', 'tlsh'], help='Enable LSH content similarity')
    
    args = parser.parse_args()
    
    signal.signal(signal.SIGINT, signal_handler)
    
    # Clear screen
    if os.name == 'posix':
        os.system('clear')
    else:
        os.system('cls')
    
    if not args.quiet:
        print_banner()
    
    print(f"\n{Colors.CYAN}[*] Target Domain: {args.domain}{Colors.END}")
    print(f"{Colors.CYAN}[*] Threads: {args.threads}{Colors.END}")
    print(f"{Colors.CYAN}[*] 100+ Permutation Techniques: ENABLED{Colors.END}")
    print(f"{Colors.CYAN}[*] Generating permutations...{Colors.END}")
    
    options = {
        'geoip': not args.no_geoip,
        'banners': not args.no_banner,
        'mxcheck': not args.no_mxcheck,
        'whois': not args.no_whois,
        'phash': args.phash,
        'lsh': args.lsh
    }
    
    session = Session(args.domain, options)
    session.scan()
    
    while session.complete < session.total:
        time.sleep(0.5)
        if not args.quiet:
            progress = int((session.complete / session.total) * 100)
            bar = '█' * (progress // 2) + '░' * (50 - progress // 2)
            sys.stdout.write(f'\r{Colors.CYAN}[{bar}]{Colors.END} {session.complete}/{session.total} domains')
            sys.stdout.flush()
    
    print()
    
    summary = session.get_summary()
    if not args.quiet:
        print_summary(summary, args.domain)
        print_results_table(session.results, show_all=args.all)
    
    if args.output:
        export_results(session.results, args.domain, args.output, args.format)
    
    if not args.quiet:
        print(f"\n{Colors.GREEN}[✓] Scan complete!{Colors.END}")
        print(f"{Colors.GRAY}Found {summary['registered']} registered domains out of {summary['total']} permutations{Colors.END}")

if __name__ == '__main__':
    main()
